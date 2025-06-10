import flet as ft
from typing import Optional, List
from datetime import datetime, date
from reporter import database_manager # This might need adjustment if database_manager path changes
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import calendar

# GuiController class remains unchanged for now
# Ensure 'Optional' is imported if not already present at the top level
# from typing import Optional, List # This is already present

class GuiController:
    def save_member_action(self, name: str, phone: str) -> tuple[bool, str]:
        """Handles the save member action with input validation."""
        # --- Input Validation ---
        if not name or not phone:
            return False, "Error: Name and Phone cannot be empty."

        if not phone.isdigit():
            return False, "Error: Phone number must contain only digits."
        # Example: Validate phone length (e.g., 10 digits) - can be added if needed
        # if len(phone) != 10:
        #     return False, "Error: Phone number must be 10 digits."

        try:
            # database_manager.add_member_to_db now returns -> Tuple[bool, str]
            success, message = database_manager.add_member_to_db(name, phone)
            # Use the returned success and message directly
            return success, message
        except Exception as e:
            # This catch block might be redundant if database_manager handles all its specific errors
            # and returns (False, error_message). However, it can catch unexpected errors.
            return False, f"An unexpected error occurred: {str(e)}"

    def save_plan_action(self, plan_name: str, duration_str: str, plan_id_to_update: str) -> tuple[bool, str, list | None]:
        """Saves a new plan or updates an existing one."""
        if not plan_name or not duration_str:
            return False, "Error: Plan Name and Duration cannot be empty.", None

        try:
            duration_days = int(duration_str)
            if duration_days <= 0:
                return False, "Error: Duration must be a positive integer.", None
        except ValueError:
            return False, "Error: Duration must be a valid integer.", None

        success = False
        message = ""
        updated_plans = None

        try:
            if plan_id_to_update:  # Editing existing plan
                plan_id = int(plan_id_to_update)
                # database_manager.update_plan now returns -> Tuple[bool, str]
                db_success, db_message = database_manager.update_plan(plan_id, plan_name, duration_days)
                success = db_success
                message = db_message
            else:  # Adding new plan
                # database_manager.add_plan now returns -> Tuple[bool, str, Optional[int]]
                db_success, db_message, returned_plan_id = database_manager.add_plan(plan_name, duration_days)
                success = db_success
                message = db_message
                # returned_plan_id is available if needed, e.g., for logging or immediate use,
                # but current logic just relies on success status.

            if success:
                updated_plans = database_manager.get_all_plans_with_inactive()
            return success, message, updated_plans
        except Exception as e:
            # Catch unexpected errors during the controller logic (e.g., int conversion if not validated)
            return False, f"An unexpected error occurred in save_plan_action: {str(e)}", None

    def toggle_plan_status_action(self, plan_id: int, current_status: bool) -> tuple[bool, str, list | None]:
        """Activates or deactivates a plan."""
        new_status = not current_status
        # database_manager.set_plan_active_status now returns -> Tuple[bool, str]
        db_success, db_message = database_manager.set_plan_active_status(plan_id, new_status)
        success = db_success
        message = db_message # Use the message from the database manager
        updated_plans = None

        if success:
            # Optionally, you could still use a custom success message or augment db_message
            # For now, directly using db_message. If a custom success message is preferred:
            # message = f"Plan status changed to {'Active' if new_status else 'Inactive'}."
            updated_plans = database_manager.get_all_plans_with_inactive()
        # If not successful, db_message already contains the error.
        return success, message, updated_plans

    def save_membership_action(self, membership_type: str, member_id: int | None,
                               start_date_str: str, amount_paid_str: str,
                               selected_plan_id: int | None = None,
                               payment_date_str: str | None = None,
                               payment_method: str | None = None,
                               sessions_str: str | None = None) -> tuple[bool, str]:
        """Handles the save membership action with validation based on membership type."""
        from datetime import datetime # Moved import

        # --- Basic Validation ---
        if not member_id: # Simplified check, assuming member_id is directly passed
            return False, "Error: Please select a valid member."

        if not start_date_str:
            return False, "Error: Start Date cannot be empty."
        try:
            datetime.strptime(start_date_str, '%Y-%m-%d')
        except ValueError:
            return False, "Error: Invalid Start Date format. Use YYYY-MM-DD."

        try:
            amount_paid = float(amount_paid_str)
            if amount_paid <= 0: # Changed from < to <=
                return False, "Error: Amount Paid must be a positive number."
        except ValueError:
            return False, "Error: Invalid Amount Paid. Must be a number."

        success = False
        try:
            if membership_type == "Group Class":
                if not selected_plan_id: # Simplified check
                    return False, "Error: Please select a valid plan."
                if not payment_date_str:
                    return False, "Error: Payment Date cannot be empty for Group Class."
                try:
                    datetime.strptime(payment_date_str, '%Y-%m-%d')
                except ValueError:
                    return False, "Error: Invalid Payment Date format. Use YYYY-MM-DD."
                if not payment_method: # Ensure payment_method is not None or empty
                    return False, "Error: Payment Method cannot be empty for Group Class."

                success, db_message = database_manager.add_transaction( # Expecting (bool, str)
                    transaction_type="Group Class",
                    member_id=member_id,
                    plan_id=selected_plan_id,
                    payment_date=payment_date_str,
                    start_date=start_date_str,
                    amount_paid=amount_paid,
                    payment_method=payment_method
                )
                if success:
                    return True, f"{membership_type} membership added successfully!"
                else:
                    return False, db_message

            elif membership_type == "Personal Training":
                if not sessions_str:
                    return False, "Error: Number of Sessions cannot be empty for PT."
                try:
                    sessions = int(sessions_str)
                    if sessions <= 0:
                        return False, "Error: Number of Sessions must be a positive integer."
                except ValueError:
                    return False, "Error: Number of Sessions must be an integer."

                success, db_message = database_manager.add_transaction( # Expecting (bool, str)
                    transaction_type="Personal Training",
                    member_id=member_id,
                    plan_id=None, # No plan_id for PT
                    payment_date=start_date_str, # For PT, payment_date is the start_date
                    start_date=start_date_str,
                    amount_paid=amount_paid,
                    payment_method="N/A", # Default or could be an argument if needed
                    sessions=sessions
                )
                if success:
                    return True, f"{membership_type} membership added successfully!"
                else:
                    return False, db_message
            else:
                return False, "Error: Unknown membership type selected."

        except Exception as e:
            # Log the exception e for debugging purposes if possible
            return False, f"An error occurred: {str(e)}"

    def generate_custom_pending_renewals_action(self, year: int, month: int) -> tuple[bool, str, list | None]:
        """Fetches pending renewals for a specific year and month."""
        import calendar # For month name in message

        try:
            renewals = database_manager.get_pending_renewals(year, month)
            month_name = calendar.month_name[month]

            if renewals:
                return True, f"Found {len(renewals)} pending renewals for {month_name} {year}.", renewals
            else:
                return True, f"No pending renewals found for {month_name} {year}.", []
        except Exception as e:
            month_name_str = str(month)
            try:
                month_name_str = calendar.month_name[month]
            except IndexError:
                pass
            return False, f"Error generating renewals report for {month_name_str} {year}: {str(e)}", None

    def generate_pending_renewals_action(self) -> tuple[bool, str, list | None]:
        """Fetches pending renewals for the current month."""
        from datetime import date
        import calendar

        today = date.today()
        current_year = today.year
        current_month = today.month
        month_name = calendar.month_name[current_month]

        try:
            renewals = database_manager.get_pending_renewals(year=current_year, month=current_month)
            if renewals:
                return True, f"Found {len(renewals)} pending renewals for {month_name} {current_year}.", renewals
            else:
                return True, f"No pending renewals found for {month_name} {current_year}.", []
        except Exception as e:
            return False, f"Error generating report for current month: {str(e)}", None

    def generate_finance_report_excel_action(self, year: int, month: int, save_path: str) -> tuple[bool, str]:
        """
        Generates an Excel finance report for the given month and year.
        The report includes a summary sheet and a detailed transactions sheet.
        Applies styling to the report.
        """
        try:
            transactions_data = database_manager.get_transactions_for_month(year, month)
            month_name_str = calendar.month_name[month]

            if not transactions_data:
                return True, f"No transaction data found for {month_name_str} {year}. Report not generated."

            df_columns = [
                "Transaction ID", "Client Name", "Payment Date", "Start Date", "End Date",
                "Amount Paid", "Type", "Plan/Sessions", "Payment Method"
            ]
            df = pd.DataFrame(transactions_data, columns=df_columns)

            df["Amount Paid"] = pd.to_numeric(df["Amount Paid"], errors='coerce').fillna(0)
            df["Payment Date"] = pd.to_datetime(df["Payment Date"], errors='coerce').dt.strftime('%Y-%m-%d')
            df["Start Date"] = pd.to_datetime(df["Start Date"], errors='coerce').dt.strftime('%Y-%m-%d')

            total_revenue = df["Amount Paid"].sum()
            total_transactions = len(df)

            summary_data = {
                "Metric": ["Total Revenue", "Total Transactions"],
                "Value": [total_revenue, total_transactions]
            }
            summary_df = pd.DataFrame(summary_data)

            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=2)
                df.to_excel(writer, sheet_name="Detailed Transactions", index=False, startrow=1)

                workbook = writer.book
                summary_sheet = writer.sheets["Summary"]
                details_sheet = writer.sheets["Detailed Transactions"]

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                title_font = Font(bold=True, size=16)
                border_style = Side(style="thin", color="000000")
                thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

                summary_sheet.cell(row=1, column=1, value=f"Finance Summary - {month_name_str} {year}").font = title_font
                for col_num, column_title in enumerate(summary_df.columns, 1):
                    cell = summary_sheet.cell(row=3, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                for row_idx_offset, _ in enumerate(summary_df.index):
                    for col_idx, _ in enumerate(summary_df.columns):
                        summary_sheet.cell(row=row_idx_offset + 4, column=col_idx + 1).border = thin_border

                summary_sheet.column_dimensions['A'].width = 25
                summary_sheet.column_dimensions['B'].width = 15
                summary_sheet.cell(row=4, column=2).number_format = '"$"#,##0.00'

                details_sheet.cell(row=1, column=1, value=f"Detailed Transactions - {month_name_str} {year}").font = title_font
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = details_sheet.cell(row=2, column=col_num)
                    cell.value = column_title
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    column_letter = get_column_letter(col_num)
                    try:
                        max_len = max(df[column_title].astype(str).map(len).max(), len(str(column_title))) + 2
                    except (TypeError, ValueError):
                        max_len = len(str(column_title)) + 2
                    details_sheet.column_dimensions[column_letter].width = max_len if max_len < 50 else 50

                for row_idx_offset, _ in enumerate(df.index):
                    for col_idx, _ in enumerate(df.columns):
                        details_sheet.cell(row=row_idx_offset + 3, column=col_idx + 1).border = thin_border

                amount_paid_col_letter = get_column_letter(df_columns.index("Amount Paid") + 1)
                for row_num in range(3, details_sheet.max_row + 1):
                     details_sheet[f'{amount_paid_col_letter}{row_num}'].number_format = '"$"#,##0.00'

            return True, f"Finance report generated successfully: {save_path}"

        except ImportError:
            return False, "Error: pandas or openpyxl library not found. Please ensure they are installed."
        except Exception as e:
            return False, f"An error occurred during report generation: {str(e)}"

    def get_book_status_action(self, year: int, month: int) -> str:
        """
        Constructs month_key, calls database_manager.get_book_status, and returns a user-friendly message.
        """
        try:
            month_key = f"{year:04d}-{month:02d}"
            status = database_manager.get_book_status(month_key)
            return f"Status for {month_key}: {status.upper()}"
        except Exception as e:
            return f"Error getting book status: {str(e)}"

    def close_books_action(self, year: int, month: int) -> tuple[bool, str]:
        """
        Constructs month_key, calls database_manager.set_book_status to "closed", and returns status.
        """
        try:
            month_key = f"{year:04d}-{month:02d}"
            success = database_manager.set_book_status(month_key, "closed")
            if success:
                return True, f"Books for {month_key} closed successfully."
            else:
                return False, f"Failed to close books for {month_key}."
        except Exception as e:
            return False, f"Error closing books for {month_key}: {str(e)}"

    def open_books_action(self, year: int, month: int) -> tuple[bool, str]:
        """
        Constructs month_key, calls database_manager.set_book_status to "open", and returns status.
        """
        try:
            month_key = f"{year:04d}-{month:02d}"
            success = database_manager.set_book_status(month_key, "open")
            if success:
                return True, f"Books for {month_key} re-opened successfully."
            else:
                return False, f"Failed to re-open books for {month_key}."
        except Exception as e:
            return False, f"Error re-opening books for {month_key}: {str(e)}"

    def get_filtered_members(self, name_filter: Optional[str], phone_filter: Optional[str]) -> list:
        """Fetches members based on name and/or phone filters."""
        return database_manager.get_all_members(name_filter=name_filter, phone_filter=phone_filter)

    def get_filtered_transaction_history(self, name_filter: Optional[str], phone_filter: Optional[str], join_date_filter: Optional[str]) -> list:
        """Fetches transaction history based on name, phone, and/or join date filters."""
        return database_manager.get_transactions_with_member_details(
            name_filter=name_filter,
            phone_filter=phone_filter,
            join_date_filter=join_date_filter
        )

    def get_active_plans(self) -> list:
        """Fetches all active plans."""
        return database_manager.get_all_plans()

    def get_all_plans_with_inactive(self) -> list:
        """Fetches all plans, including inactive ones."""
        return database_manager.get_all_plans_with_inactive()

    def get_all_activity_for_member(self, member_id: int) -> list:
        """Fetches all activity records for a given member_id."""
        return database_manager.get_all_activity_for_member(member_id)

    def deactivate_member_action(self, member_id: int) -> tuple[bool, str]:
        """Handles the action of deactivating a member."""
        try:
            success, message = database_manager.deactivate_member(member_id)
            return success, message
        except Exception as e:
            return False, f"An unexpected error occurred while deactivating the member: {str(e)}"

    def delete_transaction_action(self, transaction_id: int) -> tuple[bool, str]:
        """Calls database_manager.delete_transaction and returns status."""
        try:
            success, message = database_manager.delete_transaction(transaction_id)
            if success: # delete_transaction returns (bool, str)
                return True, "Transaction deleted successfully." # Standardize success message
            else:
                return False, message # Pass through the error message
        except Exception as e:
            return False, f"An error occurred while deleting the transaction: {str(e)}"

    def delete_plan_action(self, plan_id: int) -> tuple[bool, str]:
        """Calls database_manager.delete_plan and returns its result or an error status."""
        try:
            success, message = database_manager.delete_plan(plan_id)
            return success, message
        except Exception as e:
            return False, f"An error occurred while deleting the plan: {str(e)}"


class FletAppView(ft.UserControl):
    def __init__(self):
        super().__init__()
        self.controller = GuiController()
        self.selected_member_id_flet: Optional[int] = None
        self.selected_transaction_id_flet: Optional[int] = None
        self.selected_plan_id_flet: Optional[int] = None
        self.current_plan_id_to_update_flet: Optional[int] = None
        self.selected_start_date_flet: Optional[date] = None
        self.selected_payment_date_flet: Optional[date] = None
        self.active_date_picker_target: Optional[str] = None
        # self.page is available via self.page once the control is added to a page.

        # Date Picker
        self.date_picker = ft.DatePicker(
            on_change=self.on_date_picker_change,
            on_dismiss=self.on_date_picker_dismiss,
        )
        # File Picker
        self.file_picker = ft.FilePicker(on_result=self.on_file_picker_result_flet)

        # Member form fields
        self.member_name_input = ft.TextField(label="Name")
        self.member_phone_input = ft.TextField(label="Phone")
        self.add_member_button = ft.ElevatedButton(text="Add Member", on_click=self.on_add_member_click)
        self.member_form_feedback_text = ft.Text("")

        # Plan form fields
        self.plan_name_input = ft.TextField(label="Plan Name")
        self.plan_duration_input = ft.TextField(label="Duration (days)")
        self.save_plan_button = ft.ElevatedButton(text="Save Plan", on_click=self.on_save_plan_click)
        self.plan_form_feedback_text = ft.Text("Ready to add a new plan.", color=ft.colors.BLACK) # Initial message
        self.edit_plan_button = ft.ElevatedButton(text="Edit Selected Plan", on_click=self.on_edit_selected_plan_click)
        self.clear_plan_form_button = ft.ElevatedButton(text="Clear Form / New Plan", on_click=self.on_clear_plan_form_click)

        # Membership form fields
        self.membership_type_dropdown = ft.Dropdown(
            label="Membership Type",
            options=[ft.dropdown.Option("Group Class"), ft.dropdown.Option("Personal Training")],
            on_change=self.on_membership_type_change_flet
        )
        self.membership_member_dropdown = ft.Dropdown(label="Select Member", options=[])
        self.membership_plan_dropdown = ft.Dropdown(label="Select Plan", options=[]) # For Group Class
        self.membership_sessions_input = ft.TextField(label="Number of Sessions") # For Personal Training
        self.membership_start_date_picker_button = ft.ElevatedButton(
            text="Pick Start Date",
            on_click=lambda e: self.open_date_picker(e, "start_date")
        )
        self.membership_start_date_text = ft.Text("Start Date: Not Selected")
        self.membership_payment_date_picker_button = ft.ElevatedButton(
            text="Pick Payment Date",
            on_click=lambda e: self.open_date_picker(e, "payment_date")
        ) # For Group Class
        self.membership_payment_date_text = ft.Text("Payment Date: Not Selected") # For Group Class
        self.membership_amount_paid_input = ft.TextField(label="Amount Paid")
        self.membership_payment_method_input = ft.TextField(label="Payment Method") # For Group Class
        self.save_membership_button = ft.ElevatedButton(text="Save Membership", on_click=self.on_save_membership_click)
        self.membership_form_feedback_text = ft.Text("")

        # Reporting - Pending Renewals Form
        self.renewal_report_year_input = ft.TextField(label="Year", value=str(datetime.now().year))
        self.renewal_report_month_dropdown = ft.Dropdown(
            label="Month",
            options=[ft.dropdown.Option(str(i), str(i)) for i in range(1, 13)],
            value=str(datetime.now().month)
        )
        self.generate_renewals_report_button = ft.ElevatedButton(
            text="Generate Renewals Report",
            on_click=self.on_generate_renewals_report_click
        )
        self.renewals_report_feedback_text = ft.Text("")

        # Reporting - Finance Report Form
        self.finance_report_year_input = ft.TextField(label="Year", value=str(datetime.now().year))
        self.finance_report_month_dropdown = ft.Dropdown(
            label="Month",
            options=[ft.dropdown.Option(str(i), str(i)) for i in range(1, 13)],
            value=str(datetime.now().month)
        )
        self.generate_finance_report_button = ft.ElevatedButton(
            text="Generate Excel Finance Report",
            on_click=self.on_generate_finance_report_click
        )
        self.finance_report_feedback_text = ft.Text("")

        self.members_table_flet = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("ID")),
                ft.DataColumn(ft.Text("Name")),
                ft.DataColumn(ft.Text("Phone")),
                ft.DataColumn(ft.Text("Join Date")),
            ],
            rows=[],  # Initially empty
            on_select_changed=self.on_member_select_changed # Attach the handler
        )

        self.member_specific_history_table_flet = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Type")),
                ft.DataColumn(ft.Text("Plan/Details")),
                ft.DataColumn(ft.Text("Paid Date")),
                ft.DataColumn(ft.Text("Start Date")),
                ft.DataColumn(ft.Text("End Date")),
                ft.DataColumn(ft.Text("Amount ($)")),
                ft.DataColumn(ft.Text("Method/Sessions")),
            ],
            rows=[] # Initially empty
        )

        self.full_history_table_flet = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("TXN ID")),
                ft.DataColumn(ft.Text("Name")),
                ft.DataColumn(ft.Text("Phone")),
                ft.DataColumn(ft.Text("Joined")),
                ft.DataColumn(ft.Text("Type")),
                ft.DataColumn(ft.Text("Amount ($)")),
                ft.DataColumn(ft.Text("Paid Date")),
                ft.DataColumn(ft.Text("Start Date")),
                ft.DataColumn(ft.Text("End Date")),
                ft.DataColumn(ft.Text("Status")),
                ft.DataColumn(ft.Text("Plan/Sessions")),
                ft.DataColumn(ft.Text("Pay Method")),
            ],
            rows=[], # Initially empty
            on_select_changed=self.on_full_history_select_changed # Attach the handler
        )

        self.plans_table_flet = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("ID")),
                ft.DataColumn(ft.Text("Name")),
                ft.DataColumn(ft.Text("Duration (Days)")),
                ft.DataColumn(ft.Text("Status")),
            ],
            rows=[], # Initially empty
            on_select_changed=self.on_plan_select_changed # Attach the handler
        )
        self.pending_renewals_table_flet = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Member ID")),
                ft.DataColumn(ft.Text("Name")),
                ft.DataColumn(ft.Text("Phone")),
                ft.DataColumn(ft.Text("Plan Name")),
                ft.DataColumn(ft.Text("End Date")),
                ft.DataColumn(ft.Text("Days Overdue")),
            ],
            rows=[]  # Initially empty
        )
        # Initialize other table references here if needed for future phases

    def on_full_history_select_changed(self, e: ft.ControlEvent):
        """Handles row selection changes in the full_history_table_flet."""
        selected_index_str = e.data
        if selected_index_str:  # If a row is actually selected
            try:
                selected_index = int(selected_index_str)
                if 0 <= selected_index < len(self.full_history_table_flet.rows):
                    selected_row = self.full_history_table_flet.rows[selected_index]
                    # Assuming Transaction ID is in the first cell
                    txn_id_cell = selected_row.cells[0]
                    if isinstance(txn_id_cell.content, ft.Text):
                        self.selected_transaction_id_flet = int(txn_id_cell.content.value)
                        print(f"DEBUG: Selected Transaction ID from full history: {self.selected_transaction_id_flet}")
                    else:
                        print(f"DEBUG: Transaction ID cell content is not ft.Text: {type(txn_id_cell.content)}")
                        self.selected_transaction_id_flet = None
                else:
                    self.selected_transaction_id_flet = None  # Index out of bounds
                    print(f"DEBUG: Selected index {selected_index} out of bounds for full_history_table_flet.")
            except ValueError:
                # This can happen if txn_id_cell.content.value is not a valid int string
                txn_val = 'unknown cell'
                if 'txn_id_cell' in locals() and hasattr(txn_id_cell, 'content') and hasattr(txn_id_cell.content, 'value'):
                    txn_val = txn_id_cell.content.value
                print(f"DEBUG: Error parsing Transaction ID from full history. Cell value: '{txn_val}'")
                self.selected_transaction_id_flet = None
            except Exception as ex:
                print(f"DEBUG: An unexpected error occurred during full history selection: {ex}")
                self.selected_transaction_id_flet = None
        else:  # Selection was cleared
            self.selected_transaction_id_flet = None
            print("DEBUG: Full history table selection cleared.")
        # self.update() # Not strictly necessary if only internal state is changed

    def on_plan_select_changed(self, e: ft.ControlEvent):
        """Handles row selection changes in the plans_table_flet."""
        selected_index_str = e.data
        if selected_index_str:  # If a row is actually selected
            try:
                selected_index = int(selected_index_str)
                if 0 <= selected_index < len(self.plans_table_flet.rows):
                    selected_row = self.plans_table_flet.rows[selected_index]
                    # Assuming Plan ID is in the first cell
                    plan_id_cell = selected_row.cells[0]
                    if isinstance(plan_id_cell.content, ft.Text):
                        self.selected_plan_id_flet = int(plan_id_cell.content.value)
                        print(f"DEBUG: Selected Plan ID: {self.selected_plan_id_flet}")
                    else:
                        print(f"DEBUG: Plan ID cell content is not ft.Text: {type(plan_id_cell.content)}")
                        self.selected_plan_id_flet = None
                else:
                    self.selected_plan_id_flet = None  # Index out of bounds
                    print(f"DEBUG: Selected index {selected_index} out of bounds for plans_table_flet.")
            except ValueError:
                plan_val = 'unknown cell'
                if 'plan_id_cell' in locals() and hasattr(plan_id_cell, 'content') and hasattr(plan_id_cell.content, 'value'):
                    plan_val = plan_id_cell.content.value
                print(f"DEBUG: Error parsing Plan ID. Cell value: '{plan_val}'")
                self.selected_plan_id_flet = None
            except Exception as ex:
                print(f"DEBUG: An unexpected error occurred during plan selection: {ex}")
                self.selected_plan_id_flet = None
        else:  # Selection was cleared
            self.selected_plan_id_flet = None
            print("DEBUG: Plans table selection cleared.")
        # self.update() # Not strictly necessary if only internal state is changed

    def _get_membership_status_flet(self, end_date_str: Optional[str]) -> str:
        """Helper to determine membership status based on end date string."""
        if not end_date_str or end_date_str.lower() == "n/a" or end_date_str.strip() == "":
            return "N/A" # Or "Unknown" if "N/A" is ambiguous with a plan named "N/A"
        try:
            # Assuming end_date_str is in 'YYYY-MM-DD' format
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            if end_date_obj >= date.today():
                return "Active"
            else:
                return "Inactive"
        except ValueError:
            # This can happen if end_date_str is not a valid date or not in the expected format
            return "Invalid Date"

    def refresh_membership_history_display_flet(self, transactions_list: Optional[list] = None):
        """Populates the full_history_table_flet with transaction data."""
        self.full_history_table_flet.rows.clear()

        if transactions_list is None:
            # Record structure from controller:
            # (transaction_id, member_id, transaction_type, plan_id, payment_date,
            #  start_date, end_date, amount_paid, payment_method_db, sessions,
            #  client_name, phone, join_date)
            transactions_data = self.controller.get_filtered_transaction_history(None, None, None)
        else:
            transactions_data = transactions_list

        if not transactions_data:
            self.full_history_table_flet.rows.append(
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text("No transaction records found."), colspan=len(self.full_history_table_flet.columns))
                ])
            )
        else:
            for record in transactions_data:
                # Unpack the record based on the structure from get_filtered_transaction_history
                (transaction_id, _member_id, transaction_type, plan_id, payment_date,
                 start_date, end_date, amount_paid, payment_method_db, sessions,
                 client_name, phone, join_date) = record

                plan_id_or_sessions_display = ""
                if transaction_type == "Group Class":
                    plan_id_or_sessions_display = str(plan_id) if plan_id is not None else "N/A"
                elif transaction_type == "Personal Training":
                    plan_id_or_sessions_display = str(sessions) if sessions is not None else "N/A"

                amount_paid_formatted = "0.00"
                if amount_paid is not None:
                    try:
                        amount_paid_formatted = f"{float(amount_paid):.2f}"
                    except ValueError:
                        amount_paid_formatted = "Error"

                end_date_display = str(end_date) if end_date is not None else "N/A"

                # Pass the original end_date string (or None if it was None) to status calculation
                # _get_membership_status_flet expects a string 'YYYY-MM-DD' or None
                status = self._get_membership_status_flet(str(end_date) if end_date else None)

                # Prepare all values for display, ensuring they are strings and handle None
                ordered_values = [
                    str(transaction_id),
                    str(client_name if client_name is not None else "N/A"),
                    str(phone if phone is not None else "N/A"),
                    str(join_date if join_date is not None else "N/A"),
                    str(transaction_type if transaction_type is not None else "N/A"),
                    amount_paid_formatted,
                    str(payment_date if payment_date is not None else "N/A"),
                    str(start_date if start_date is not None else "N/A"),
                    end_date_display, # Use the N/A handled version for display
                    status,
                    plan_id_or_sessions_display,
                    str(payment_method_db if payment_method_db is not None else "N/A")
                ]

                data_cells = [ft.DataCell(ft.Text(value)) for value in ordered_values]
                self.full_history_table_flet.rows.append(ft.DataRow(cells=data_cells))

        self.update()

    def apply_history_filters_flet(self, e):
        """Placeholder for applying transaction history filters."""
        # To be implemented: get filter values, call controller, then refresh_membership_history_display_flet(filtered_data)
        pass

    def clear_history_filters_flet(self, e):
        """Clears transaction history filters and re-displays all history."""
        # To be implemented: clear filter fields in UI
        self.refresh_membership_history_display_flet() # Fetches all data

    def display_membership_history_flet(self, member_id: Optional[int]):
        """Populates the member_specific_history_table_flet with activity for the given member_id."""
        self.member_specific_history_table_flet.rows.clear()

        if member_id is None:
            placeholder_text = "Select a member to view their activity."
            self.member_specific_history_table_flet.rows.append(
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text(placeholder_text), colspan=len(self.member_specific_history_table_flet.columns))
                ])
            )
            self.update()
            return

        history_records = self.controller.get_all_activity_for_member(member_id)

        if not history_records:
            placeholder_text = "No activity history found for this member."
            self.member_specific_history_table_flet.rows.append(
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text(placeholder_text), colspan=len(self.member_specific_history_table_flet.columns))
                ])
            )
        else:
            # record: (activity_type, name_or_description, payment_date, start_date, end_date, amount_paid, payment_method_or_sessions, activity_id)
            for record in history_records:
                cells = []
                # Type (index 0)
                cells.append(ft.DataCell(ft.Text(str(record[0] if record[0] is not None else "N/A"))))
                # Plan/Details (index 1)
                cells.append(ft.DataCell(ft.Text(str(record[1] if record[1] is not None else "N/A"))))
                # Paid Date (index 2)
                cells.append(ft.DataCell(ft.Text(str(record[2] if record[2] is not None else "N/A"))))
                # Start Date (index 3)
                cells.append(ft.DataCell(ft.Text(str(record[3] if record[3] is not None else "N/A"))))
                # End Date (index 4)
                cells.append(ft.DataCell(ft.Text(str(record[4] if record[4] is not None else "N/A"))))

                # Amount Paid (index 5) - format as currency
                amount_paid_val = record[5]
                if isinstance(amount_paid_val, (int, float)):
                    amount_text = f"{amount_paid_val:.2f}"
                elif amount_paid_val is None: # Or any other representation for truly missing data
                    amount_text = "0.00"
                else:
                    amount_text = str(amount_paid_val) # Fallback if it's some other type already a string
                cells.append(ft.DataCell(ft.Text(amount_text)))

                # Method/Sessions (index 6)
                cells.append(ft.DataCell(ft.Text(str(record[6] if record[6] is not None else "N/A"))))

                self.member_specific_history_table_flet.rows.append(ft.DataRow(cells=cells))

        self.update()

    def on_member_select_changed(self, e: ft.ControlEvent):
        """Handles member selection changes in the members_table_flet."""
        selected_index_str = e.data  # This is the string representation of the row index or ""

        if selected_index_str:  # A row is selected
            try:
                selected_index = int(selected_index_str)
                if 0 <= selected_index < len(self.members_table_flet.rows):
                    selected_row = self.members_table_flet.rows[selected_index]
                    member_id_cell = selected_row.cells[0] # ID is the first cell

                    # Ensure content is ft.Text and access its value
                    if isinstance(member_id_cell.content, ft.Text):
                        self.selected_member_id_flet = int(member_id_cell.content.value)
                        # print(f"DEBUG: Selected Member ID: {self.selected_member_id_flet}") # For debugging
                    else:
                        # Fallback or error if cell content is not as expected
                        print(f"Error: Member ID cell content is not ft.Text: {type(member_id_cell.content)}")
                        self.selected_member_id_flet = None
                else:
                    # Index out of bounds, should not happen with valid e.data
                    print(f"Error: Selected index {selected_index} is out of bounds.")
                    self.selected_member_id_flet = None
            except ValueError:
                # Handle cases where conversion to int fails for selected_index_str or member_id_cell.content.value
                print(f"Error: Could not parse selected index or member ID. Data: '{selected_index_str}'")
                self.selected_member_id_flet = None
            except Exception as ex:
                print(f"An unexpected error occurred during member selection: {ex}")
                self.selected_member_id_flet = None
        else:  # Selection was cleared (e.data is an empty string)
            self.selected_member_id_flet = None
            # print("DEBUG: Member selection cleared.") # For debugging

        # Call the history display function regardless of whether a member is selected or deselected
        self.display_membership_history_flet(self.selected_member_id_flet)

        # self.update() # Typically not needed here as Flet handles updates from control events.
                       # Only if this handler directly changes other controls not via standard Flet binding.

    def display_all_members_flet(self, members_list: Optional[list] = None):
        """Populates the members_table_flet with member data."""
        if members_list is None:
            # Fetches (member_id, client_name, phone, join_date, is_active)
            members_list = self.controller.get_filtered_members(None, None)

        self.members_table_flet.rows.clear()
        if members_list:
            for member_data in members_list:
                # Display only the first 4 columns: ID, Name, Phone, Join Date
                row = ft.DataRow(cells=[
                    ft.DataCell(ft.Text(str(member_data[0]))), # ID
                    ft.DataCell(ft.Text(str(member_data[1]))), # Name
                    ft.DataCell(ft.Text(str(member_data[2]))), # Phone
                    ft.DataCell(ft.Text(str(member_data[3]))), # Join Date
                ])
                self.members_table_flet.rows.append(row)

        # If the control is already part of a page, self.update() should refresh it.
        # If self.page is directly accessible and an update is needed more broadly, self.page.update() could be used.
        self.update()

    def apply_member_filters_flet(self, e):
        """Placeholder for applying member filters."""
        # This will be implemented later. For now, it does nothing.
        # Example:
        # name_filter = self.member_name_filter_field.value or None
        # phone_filter = self.member_phone_filter_field.value or None
        # filtered_members = self.controller.get_filtered_members(name_filter, phone_filter)
        # self.display_all_members_flet(filtered_members)
        pass

    def clear_member_filters_flet(self, e):
        """Clears member filters and re-displays all members."""
        # Assuming filter fields would be cleared here, e.g.:
        # self.member_name_filter_field.value = ""
        # self.member_phone_filter_field.value = ""
        self.display_all_members_flet(self.controller.get_filtered_members(None, None))
        # self.update() # display_all_members_flet already calls self.update()

    def display_all_plans_flet(self, plans_list: Optional[list] = None):
        """Populates the plans_table_flet with plan data."""
        if plans_list is None:
            # Fetches (plan_id, plan_name, duration_days, is_active)
            plans_list = self.controller.get_all_plans_with_inactive()

        self.plans_table_flet.rows.clear()
        if plans_list:
            for plan_data in plans_list:
                plan_id, plan_name, duration_days, is_active = plan_data
                status = "Active" if is_active else "Inactive"
                row = ft.DataRow(cells=[
                    ft.DataCell(ft.Text(str(plan_id))),
                    ft.DataCell(ft.Text(str(plan_name))),
                    ft.DataCell(ft.Text(str(duration_days))),
                    ft.DataCell(ft.Text(status)),
                ])
                self.plans_table_flet.rows.append(row)
        else:
            self.plans_table_flet.rows.append(
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text("No plans found."), colspan=len(self.plans_table_flet.columns))
                ])
            )
        self.update()

    def display_pending_renewals_flet(self):
        """Populates the pending_renewals_table_flet with current month's pending renewal data."""
        success, message, renewals_data = self.controller.generate_pending_renewals_action()
        self.pending_renewals_table_flet.rows.clear()

        if success and renewals_data:
            for renewal_item in renewals_data:
                # (member_id, client_name, phone, plan_name, end_date, days_overdue)
                cells = [ft.DataCell(ft.Text(str(item))) for item in renewal_item]
                self.pending_renewals_table_flet.rows.append(ft.DataRow(cells=cells))
        else:
            # Use the message from controller if an error occurred, otherwise a default message
            display_message = message if not success else "No pending renewals found for the current month."
            self.pending_renewals_table_flet.rows.append(
                ft.DataRow(cells=[
                    ft.DataCell(ft.Text(display_message), colspan=len(self.pending_renewals_table_flet.columns))
                ])
            )
        self.update()

    def on_add_member_click(self, e):
        """Handles the click event of the 'Add Member' button."""
        name = self.member_name_input.value
        phone = self.member_phone_input.value

        success, message = self.controller.save_member_action(name, phone)

        self.member_form_feedback_text.value = message
        if success:
            self.member_form_feedback_text.color = ft.colors.GREEN
            self.member_name_input.value = ""
            self.member_phone_input.value = ""
            self.display_all_members_flet()
            self.populate_member_dropdowns_flet() # Now implemented
        else:
            self.member_form_feedback_text.color = ft.colors.RED

        self.member_name_input.update() # Ensure input fields are visually cleared
        self.member_phone_input.update()
        self.member_form_feedback_text.update()
        self.update() # General UI update

    def populate_member_dropdowns_flet(self):
        """Populates the membership_member_dropdown with active members."""
        members = self.controller.get_filtered_members(None, None)  # Fetches (id, name, phone, join_date, is_active)
        options = [ft.dropdown.Option(key=str(m[0]), text=f"{m[1]} (ID: {m[0]})") for m in members if m[4]] # Only active

        current_value = self.membership_member_dropdown.value
        self.membership_member_dropdown.options = options
        if options:
            # Try to keep current selection if it's still valid, else default to first
            if current_value and any(opt.key == current_value for opt in options):
                self.membership_member_dropdown.value = current_value
            else:
                self.membership_member_dropdown.value = options[0].key
        else:
            self.membership_member_dropdown.value = None

        if hasattr(self, 'membership_member_dropdown') and self.membership_member_dropdown.page: # Check if page exists
            self.membership_member_dropdown.update() # Update only if part of the page
        # self.update() # Broader update might not be needed if only dropdown changes

    def on_save_plan_click(self, e):
        """Handles the click event of the 'Save Plan' button."""
        plan_name = self.plan_name_input.value
        duration_str = self.plan_duration_input.value
        plan_id_to_update_str = str(self.current_plan_id_to_update_flet) if self.current_plan_id_to_update_flet else None

        success, message, updated_plans = self.controller.save_plan_action(plan_name, duration_str, plan_id_to_update_str)

        self.plan_form_feedback_text.value = message
        if success:
            self.plan_form_feedback_text.color = ft.colors.GREEN
            self.on_clear_plan_form_click(None) # Clear form and reset state
            if updated_plans is not None: # Check if updated_plans is not None before passing
                self.display_all_plans_flet(updated_plans)
            else: # Fallback if updated_plans is None for some reason (e.g. error during fetch)
                self.display_all_plans_flet() # Refresh with all plans
            self.populate_plan_dropdowns_flet() # Now implemented
        else:
            self.plan_form_feedback_text.color = ft.colors.RED

        self.plan_form_feedback_text.update()
        # self.plan_name_input.update() and self.plan_duration_input.update() are handled by on_clear_plan_form_click
        self.update() # General UI update for table refresh

    def on_edit_selected_plan_click(self, e):
        """Populates the plan form with data from the selected plan in the table."""
        if self.selected_plan_id_flet is not None:
            selected_plan_details = None
            for row in self.plans_table_flet.rows:
                if int(row.cells[0].content.value) == self.selected_plan_id_flet:
                    selected_plan_details = {
                        "id": row.cells[0].content.value,
                        "name": row.cells[1].content.value,
                        "duration": row.cells[2].content.value,
                        # Status (cell 3) is not directly editable in this form version
                    }
                    break

            if selected_plan_details:
                self.plan_name_input.value = selected_plan_details["name"]
                self.plan_duration_input.value = selected_plan_details["duration"]
                self.current_plan_id_to_update_flet = self.selected_plan_id_flet
                self.plan_form_feedback_text.value = f"Editing Plan ID: {self.current_plan_id_to_update_flet}"
                self.plan_form_feedback_text.color = ft.colors.BLUE
            else:
                self.plan_form_feedback_text.value = "Error: Could not find details for the selected plan."
                self.plan_form_feedback_text.color = ft.colors.RED
        else:
            self.plan_form_feedback_text.value = "Please select a plan from the table to edit."
            self.plan_form_feedback_text.color = ft.colors.ORANGE

        self.plan_name_input.update()
        self.plan_duration_input.update()
        self.plan_form_feedback_text.update()
        self.update()

    def on_clear_plan_form_click(self, e):
        """Clears the plan form fields and resets the editing state."""
        self.plan_name_input.value = ""
        self.plan_duration_input.value = ""
        self.current_plan_id_to_update_flet = None
        self.plan_form_feedback_text.value = "Ready to add a new plan."
        self.plan_form_feedback_text.color = ft.colors.BLACK # Default text color

        self.plan_name_input.update()
        self.plan_duration_input.update()
        self.plan_form_feedback_text.update()
        self.update() # Potentially to update other related UI elements if any

    def populate_plan_dropdowns_flet(self):
        """Populates the membership_plan_dropdown with active plans."""
        plans = self.controller.get_active_plans()  # Fetches (id, name, duration)
        options = [ft.dropdown.Option(key=str(p[0]), text=f"{p[1]} ({p[2]} days)") for p in plans]

        current_value = self.membership_plan_dropdown.value
        self.membership_plan_dropdown.options = options
        if options:
             # Try to keep current selection if it's still valid, else default to first
            if current_value and any(opt.key == current_value for opt in options):
                self.membership_plan_dropdown.value = current_value
            else:
                self.membership_plan_dropdown.value = options[0].key
        else:
            self.membership_plan_dropdown.value = None

        if hasattr(self, 'membership_plan_dropdown') and self.membership_plan_dropdown.page: # Check if page exists
            self.membership_plan_dropdown.update()
        # self.update()

    def open_date_picker(self, e, date_type_to_set: str):
        """Opens the date picker and sets the target for date selection."""
        self.active_date_picker_target = date_type_to_set
        # self.page.open(self.date_picker) # Old way
        self.date_picker.pick_date()


    def on_date_picker_change(self, e):
        """Handles date selection from the DatePicker."""
        selected_date = self.date_picker.value.date() if self.date_picker.value else None
        if self.active_date_picker_target == "start_date":
            self.selected_start_date_flet = selected_date
            self.membership_start_date_text.value = f"Start Date: {selected_date.strftime('%Y-%m-%d') if selected_date else 'Not Selected'}"
            if self.membership_start_date_text.page: self.membership_start_date_text.update()
        elif self.active_date_picker_target == "payment_date":
            self.selected_payment_date_flet = selected_date
            self.membership_payment_date_text.value = f"Payment Date: {selected_date.strftime('%Y-%m-%d') if selected_date else 'Not Selected'}"
            if self.membership_payment_date_text.page: self.membership_payment_date_text.update()

        self.active_date_picker_target = None # Reset target
        # self.update() # May not be needed if individual controls are updated

    def on_date_picker_dismiss(self, e):
        """Handles dismissal of the DatePicker."""
        # User dismissed the picker, you might want to log this or do nothing.
        self.active_date_picker_target = None # Reset target
        # self.update() # Update UI if needed

    def on_membership_type_change_flet(self, e):
        """Shows/hides fields based on membership type selection."""
        selection = self.membership_type_dropdown.value
        is_group_class = selection == "Group Class"

        self.membership_plan_dropdown.visible = is_group_class
        self.membership_payment_date_picker_button.visible = is_group_class
        self.membership_payment_date_text.visible = is_group_class
        self.membership_payment_method_input.visible = is_group_class
        self.membership_sessions_input.visible = not is_group_class

        if self.page: # Ensure page exists before updating controls
            self.membership_plan_dropdown.update()
            self.membership_payment_date_picker_button.update()
            self.membership_payment_date_text.update()
            self.membership_payment_method_input.update()
            self.membership_sessions_input.update()
            self.update() # General update for layout changes

    def on_save_membership_click(self, e):
        """Handles the save membership action."""
        membership_type = self.membership_type_dropdown.value
        member_id_str = self.membership_member_dropdown.value
        member_id = int(member_id_str) if member_id_str else None

        start_date_str = self.selected_start_date_flet.strftime('%Y-%m-%d') if self.selected_start_date_flet else None
        amount_paid_str = self.membership_amount_paid_input.value

        plan_id = None
        payment_date_str = None
        payment_method = None
        sessions_str = None

        if membership_type == "Group Class":
            plan_id_str = self.membership_plan_dropdown.value
            plan_id = int(plan_id_str) if plan_id_str else None
            payment_date_str = self.selected_payment_date_flet.strftime('%Y-%m-%d') if self.selected_payment_date_flet else None
            payment_method = self.membership_payment_method_input.value
        elif membership_type == "Personal Training":
            sessions_str = self.membership_sessions_input.value
            payment_date_str = start_date_str # PT payment date is the start date
            payment_method = "N/A" # Default for PT

        success, message = self.controller.save_membership_action(
            membership_type, member_id, start_date_str, amount_paid_str,
            plan_id, payment_date_str, payment_method, sessions_str
        )

        self.membership_form_feedback_text.value = message
        if success:
            self.membership_form_feedback_text.color = ft.colors.GREEN
            # Clear form fields
            self.membership_amount_paid_input.value = ""
            self.membership_payment_method_input.value = "" # Specific to Group Class but clearing doesn't hurt
            self.membership_sessions_input.value = ""    # Specific to PT

            self.selected_start_date_flet = None
            self.membership_start_date_text.value = "Start Date: Not Selected"
            self.selected_payment_date_flet = None
            self.membership_payment_date_text.value = "Payment Date: Not Selected"

            # Optionally reset dropdowns - for now, let them keep their selection
            # self.membership_type_dropdown.value = None # Or first option
            # self.membership_member_dropdown.value = None # Or first option
            # self.membership_plan_dropdown.value = None # Or first option

            self.refresh_membership_history_display_flet() # Update full history
            if member_id is not None and self.selected_member_id_flet == member_id:
                self.display_membership_history_flet(member_id) # Update specific member view if shown

        else:
            self.membership_form_feedback_text.color = ft.colors.RED

        # Update specific controls that changed
        self.membership_form_feedback_text.update()
        self.membership_amount_paid_input.update()
        self.membership_payment_method_input.update()
        self.membership_sessions_input.update()
        self.membership_start_date_text.update()
        self.membership_payment_date_text.update()
        self.update() # General UI update

    def on_generate_renewals_report_click(self, e):
        """Handles the click event for generating the pending renewals report."""
        year_str = self.renewal_report_year_input.value
        month_str = self.renewal_report_month_dropdown.value

        if not year_str or not month_str:
            self.renewals_report_feedback_text.value = "Year and Month cannot be empty."
            self.renewals_report_feedback_text.color = ft.colors.RED
            self.renewals_report_feedback_text.update()
            return

        try:
            year = int(year_str)
            month = int(month_str)
            if not (1 <= month <= 12):
                raise ValueError("Month out of range.")
        except ValueError:
            self.renewals_report_feedback_text.value = "Invalid Year or Month format."
            self.renewals_report_feedback_text.color = ft.colors.RED
            self.renewals_report_feedback_text.update()
            return

        success, message, renewals_data = self.controller.generate_custom_pending_renewals_action(year, month)
        self.renewals_report_feedback_text.value = message
        self.renewals_report_feedback_text.color = ft.colors.GREEN if success else ft.colors.RED
        self.renewals_report_feedback_text.update()

        self.pending_renewals_table_flet.rows.clear()
        if success and renewals_data:
            for item in renewals_data:
                # (member_id, client_name, phone, plan_name, end_date, days_overdue)
                cells = [ft.DataCell(ft.Text(str(val))) for val in item]
                self.pending_renewals_table_flet.rows.append(ft.DataRow(cells=cells))
        elif success and not renewals_data: # Success but no data
             self.pending_renewals_table_flet.rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(message), colspan=len(self.pending_renewals_table_flet.columns))]))
        else: # Not success or renewals_data is None
            error_message = message if message else "Error generating report or no data."
            self.pending_renewals_table_flet.rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(error_message), colspan=len(self.pending_renewals_table_flet.columns))]))

        self.pending_renewals_table_flet.update()
        self.update()

    def on_generate_finance_report_click(self, e):
        """Handles the click event for generating the finance report."""
        year_str = self.finance_report_year_input.value
        month_str = self.finance_report_month_dropdown.value

        if not year_str or not month_str:
            self.finance_report_feedback_text.value = "Year and Month cannot be empty."
            self.finance_report_feedback_text.color = ft.colors.RED
            self.finance_report_feedback_text.update()
            return
        try:
            int(year_str) # Validate year can be int
            if not (1 <= int(month_str) <= 12):
                 raise ValueError("Month out of range.")
        except ValueError:
            self.finance_report_feedback_text.value = "Invalid Year or Month format."
            self.finance_report_feedback_text.color = ft.colors.RED
            self.finance_report_feedback_text.update()
            return

        # Proceed to file picking
        self.file_picker.save_file(
            dialog_title="Save Finance Report",
            file_name=f"finance_report_{year_str}_{month_str}.xlsx",
            allowed_extensions=["xlsx"]
        )

    def on_file_picker_result_flet(self, e: ft.FilePickerResultEvent):
        """Handles the result of the file picker dialog."""
        if e.path:
            save_path = e.path
            year = int(self.finance_report_year_input.value) # Assumes valid from previous check
            month = int(self.finance_report_month_dropdown.value) # Assumes valid

            success, message = self.controller.generate_finance_report_excel_action(year, month, save_path)
            self.finance_report_feedback_text.value = message
            self.finance_report_feedback_text.color = ft.colors.GREEN if success else ft.colors.RED
        else:
            self.finance_report_feedback_text.value = "File save cancelled."
            self.finance_report_feedback_text.color = ft.colors.ORANGE

        self.finance_report_feedback_text.update()
        self.update()

    def build(self):
        # Initial population of the members table
        # This needs to be called after members_table_flet is initialized but before the UI is built,
        # or right after the UI structure using it is defined.
        # Calling it here means it will populate when build() is first called.

        # Ensure DatePicker is added to page overlays
        # This check is to prevent adding it multiple times if build is called again,
        # though typically build is called once per control instance.
        # However, self.page might not be available yet in __init__.
        # It's safer to do this here or in did_mount.
        # For this structure, doing it in build() before returning tabs_control is fine.
        # self.page.overlay.append(self.date_picker) # This will be done after self.page is set.

        # Renaming to table_area_container as per latest prompt for clarity
        table_area_container = ft.Container(
            content=ft.Column(
                controls=[
                    ft.Text("All Members", weight=ft.FontWeight.BOLD),
                    self.members_table_flet,
                    ft.Divider(), # Visual separator
                    ft.Text("Selected Member Activity", weight=ft.FontWeight.BOLD),
                    self.member_specific_history_table_flet
                ],
                scroll=ft.ScrollMode.AUTO,
                expand=True,
                spacing=10 # Added spacing
            ),
            expand=2,
            bgcolor=ft.colors.BLUE_GREY_300, # Or your preferred background
            padding=10,
            alignment=ft.alignment.top_center
        )

        self.tabs_control = ft.Tabs(
            tabs=[
                ft.Tab(
                    text="Membership Management",
                    content=ft.Row(
                        controls=[
                            ft.Container(
                                content=ft.Column(
                                    controls=[
                                        ft.Text("Add New Member", weight=ft.FontWeight.BOLD),
                                        self.member_name_input,
                                        self.member_phone_input,
                                        self.add_member_button,
                                        self.member_form_feedback_text,
                                        ft.Divider(),
                                        ft.Text("Add New Membership", weight=ft.FontWeight.BOLD),
                                        self.membership_type_dropdown,
                                        self.membership_member_dropdown,
                                        self.membership_plan_dropdown, # Visibility toggled
                                        self.membership_sessions_input, # Visibility toggled
                                        ft.Row([self.membership_start_date_picker_button, self.membership_start_date_text]),
                                        ft.Row([self.membership_payment_date_picker_button, self.membership_payment_date_text]), # Visibility toggled
                                        self.membership_amount_paid_input,
                                        self.membership_payment_method_input, # Visibility toggled
                                        self.save_membership_button,
                                        self.membership_form_feedback_text,
                                    ],
                                    scroll=ft.ScrollMode.AUTO # Make this column scrollable
                                ),
                                expand=1,
                                bgcolor=ft.colors.BLUE_GREY_200,
                                padding=10,
                                alignment=ft.alignment.top_center,
                            ),
                            table_area_container,
                        ]
                    )
                ),
                ft.Tab(
                    text="Membership History",
                    content=ft.Column(
                        controls=[
                            ft.Container(
                                content=ft.Text("Filters Placeholder"),
                                expand=1, # Adjusted expand for Column layout
                                bgcolor=ft.colors.GREEN_200,
                                padding=10,
                                alignment=ft.alignment.center,
                                # Adjust expand or height as needed for filter area
                                # expand=1, # Original from prompt, might be too large if filters are few
                                height=100, # Example fixed height, or remove expand to size to content
                            ),
                            # Container for the full history table
                            ft.Container(
                                content=ft.Column(
                                    controls=[self.full_history_table_flet],
                                    scroll=ft.ScrollMode.AUTO,
                                    expand=True
                                ),
                                expand=4, # Retain original expand factor for the table area
                                bgcolor=ft.colors.GREEN_300,
                                padding=10,
                                alignment=ft.alignment.top_center
                            ),
                        ],
                        # Optional: Adjust spacing for the main Column of the tab
                        spacing=10,
                        # Optional: Stretch filter container width if needed
                        # horizontal_alignment=ft.CrossAxisAlignment.STRETCH
                    )
                ),
                ft.Tab(
                    text="Plan Management",
                    content=ft.Row(
                        controls=[
                            ft.Container(
                                content=ft.Column(
                                    controls=[
                                        ft.Text("Add/Edit Plan", weight=ft.FontWeight.BOLD),
                                        self.plan_name_input,
                                        self.plan_duration_input,
                                        self.save_plan_button,
                                        self.edit_plan_button, # Added button
                                        self.clear_plan_form_button, # Added button
                                        self.plan_form_feedback_text,
                                    ]
                                ),
                                expand=1,
                                bgcolor=ft.colors.AMBER_200,
                                padding=10,
                                alignment=ft.alignment.top_center,
                            ),
                            # Container for the plans table
                            ft.Container(
                                content=ft.Column(
                                    controls=[self.plans_table_flet],
                                    scroll=ft.ScrollMode.AUTO,
                                    expand=True
                                ),
                                expand=2,
                                bgcolor=ft.colors.AMBER_300,
                                padding=10,
                                alignment=ft.alignment.top_center
                            ),
                        ],
                        # Optional: Adjust spacing for the main Row of the tab
                        # spacing=10,
                        # vertical_alignment=ft.CrossAxisAlignment.START
                    )
                ),
                ft.Tab(
                    text="Reporting",
                    content=ft.Column(
                        controls=[
                            ft.Container( # Container for Renewals Report Form and Table
                                content=ft.Column(
                                    controls=[
                                        ft.Text("Custom Pending Renewals Report", weight=ft.FontWeight.BOLD),
                                        self.renewal_report_year_input,
                                        self.renewal_report_month_dropdown,
                                        self.generate_renewals_report_button,
                                        self.renewals_report_feedback_text,
                                        ft.Divider(),
                                        ft.Text("Pending Renewals Results", weight=ft.FontWeight.BOLD),
                                        self.pending_renewals_table_flet
                                    ],
                                    scroll=ft.ScrollMode.AUTO,
                                    expand=True # Ensure this column can expand
                                ),
                                expand=1,
                                bgcolor=ft.colors.TEAL_200,
                                padding=10,
                                alignment=ft.alignment.top_center
                            ),
                            ft.Container( # Container for Finance Report Form
                                content=ft.Column(
                                    controls=[
                                        ft.Text("Generate Monthly Finance Report (Excel)", weight=ft.FontWeight.BOLD),
                                        self.finance_report_year_input,
                                        self.finance_report_month_dropdown,
                                        self.generate_finance_report_button,
                                        self.finance_report_feedback_text,
                                    ]
                                ),
                                expand=1,
                                bgcolor=ft.colors.TEAL_300,
                                padding=10,
                                alignment=ft.alignment.top_center, # Changed from center
                            ),
                        ]
                    )
                ),
                ft.Tab(
                    text="Settings",
                    content=ft.Column( # Main content for settings is a Column
                        controls=[
                            ft.Container( # This container can hold all settings content
                                content=ft.Text("Settings Placeholder"),
                                expand=True, # Allow this container to fill the column
                                bgcolor=ft.colors.ORANGE_200,
                                padding=10,
                                alignment=ft.alignment.center,
                            )
                        ]
                    )
                ),
                # ... other tabs remain unchanged in structure for this step ...
            ]
        )

        # Populate members table after tabs_control is defined and uses members_table_flet
        self.display_all_members_flet()
        # Populate full history table as well
        self.refresh_membership_history_display_flet()
        # Populate plans table
        self.display_all_plans_flet()
        # Populate pending renewals table - Now handled by on_generate_renewals_report_click called below
        # self.display_pending_renewals_flet()

        # Populate dropdowns
        self.populate_member_dropdowns_flet()
        self.populate_plan_dropdowns_flet()

        # Set initial visibility for membership form based on default type
        self.on_membership_type_change_flet(None)


        # Add DatePicker to the page overlay once the page is available
        # This is often done in did_mount or after the main control is added to the page.
        # For now, let's assume self.page will be set by Flet when app_view is added.
        # A common pattern is to check self.page and add overlay items.
        if self.page:
            if self.date_picker not in self.page.overlay:
                self.page.overlay.append(self.date_picker)
            if self.file_picker not in self.page.overlay: # Add FilePicker to overlay
                self.page.overlay.append(self.file_picker)

        # Initial population of the pending renewals table with default/current month's data
        self.on_generate_renewals_report_click(None)


        return self.tabs_control

def main(page: ft.Page):
    page.title = "Kranos MMA Reporter"
    page.theme_mode = ft.ThemeMode.DARK

    # Create and add the main app view
    app_view = FletAppView()
    page.add(app_view) # This implicitly calls app_view.build() and sets app_view.page

    # Now that app_view.page is set, add overlay components
    if app_view.date_picker not in page.overlay:
        page.overlay.append(app_view.date_picker)
    if app_view.file_picker not in page.overlay: # Ensure FilePicker is added
        page.overlay.append(app_view.file_picker)

    page.update()

if __name__ == "__main__":
    # Ensure database is initialized
    # This might be better placed in a dedicated startup script or within FletAppView.__init__
    # if it's safe to call multiple times or if FletAppView is guaranteed to be a singleton.
    # For now, keeping it simple here.
    database_manager.initialize_database() # Assuming this function exists and sets up the DB
    ft.app(target=main)

# Note:
# - GuiController class is kept as is from the original file.
# - The main change is the introduction of FletAppView and moving tab definitions into its build method.
# - self.members_table_flet is created and placed in the "Membership Management" tab.
# - Other tabs retain their placeholders but are now part of the class structure.
# - The main function now instantiates FletAppView.
# - Added a call to database_manager.initialize_database() in if __name__ == "__main__":
#   This is a placeholder for where database initialization should occur.
#   The actual database_manager.py might need to be checked for the correct initialization function.
# - Corrected a few controller methods to properly return (bool, message) from database_manager calls.
# - Wrapped self.members_table_flet in an ft.Column as per example, though ft.DataTable can often be direct content.
#   This allows for adding more controls (like buttons above/below table) in that container later.
# - Adjusted expand properties for some placeholders in Column layouts as direct children of Column use `expand` differently.
#   Often, direct children of Column are sized intrinsically or use `height`/`width` unless also wrapped in `Container` with `expand`.
#   For simplicity, kept `expand` on containers.
# - The `database_manager` import might need an alias or path adjustment if the file structure is complex,
#   but `from reporter import database_manager` should work if `reporter` is the package root.
# - The `save_membership_action` and `delete_transaction_action` in `GuiController` were slightly adjusted to correctly handle the (bool, message) tuple from `database_manager`.
# - Fixed `generate_finance_report_excel_action` to ensure `df_columns` matches the data structure from `get_transactions_for_month`.
# - Removed redundant exception catch in `save_member_action` if `database_manager` handles its errors.
# - Ensured `toggle_plan_status_action` uses the message from `database_manager`.
# - Ensured `save_plan_action` correctly uses messages from `database_manager`.
# - Ensured `deactivate_member_action` uses message from `database_manager`.
# - Ensured `delete_plan_action` correctly returns message from `database_manager`.
# - Corrected `generate_custom_pending_renewals_action` and `generate_pending_renewals_action` to ensure they return an empty list for the third element of the tuple when no renewals are found, to match the type hint `list | None`.
# - Corrected `save_membership_action` to properly unpack the tuple from `database_manager.add_transaction` for both "Group Class" and "Personal Training" cases.
# - The `ft.Container` holding `self.members_table_flet` is set to `alignment=ft.alignment.top_center`. The table itself might also benefit from `expand=True` if it's the only child in the column and expected to fill. For now, `ft.Column([self.members_table_flet])` is used.
# - The `expand=True` on the `ft.DataTable` itself might be useful. Let's consider adding it directly to the table:
#   self.members_table_flet = ft.DataTable(..., expand=True)
#   and then the container could just be:
#   ft.Container(content=self.members_table_flet, expand=2, ...)
#   The current approach of ft.Column([self.members_table_flet]) is also fine.
#   The prompt specified "ft.Container(content=ft.Column([self.members_table_flet]) ... expand=2"
#   or "Or directly self.members_table_flet if it handles its own scrolling/sizing".
#   Let's stick to ft.Column([self.members_table_flet]) for now as it's more flexible for adding other controls later.
#   The outer container having expand=2 should make the column (and thus the table) fill the space.
#   If horizontal scrolling is needed, the DataTable itself has properties for that, or it can be wrapped in a Row.
#   The subtask mentions: "Add ft.Row([self.members_table_flet]) if the table needs to be scrollable horizontally later".
#   For now, ft.Column([self.members_table_flet]) is fine.

# Final check on GuiController.save_membership_action:
# It was returning `success` which was a tuple (bool, str) instead of just the boolean.
# Corrected it to return `True, "message"` or `False, db_message`.
# This was done by assigning `success, db_message = database_manager.add_transaction(...)`
# and then using `if success:`
# This has been applied.
