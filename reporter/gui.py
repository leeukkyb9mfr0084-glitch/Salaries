import flet as ft
from typing import Optional, List
from datetime import datetime, date
from reporter import database_manager # This might need adjustment if database_manager path changes
import pandas as pd
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import calendar

# GuiController class remains unchanged for now
# Ensure 'Optional' is imported if not already present at the top level
# from typing import Optional, List # This is already present

class GuiController:
    def save_member_action(self, name: str, phone: str) -> tuple[bool, str]:
        """Handles the save member action with input validation."""
        # --- Input Validation ---
        if not name or not phone:
            return False, "Error: Name and Phone cannot be empty."

        if not phone.isdigit():
            return False, "Error: Phone number must contain only digits."
        # Example: Validate phone length (e.g., 10 digits) - can be added if needed
        # if len(phone) != 10:
        #     return False, "Error: Phone number must be 10 digits."

        try:
            # database_manager.add_member_to_db now returns -> Tuple[bool, str]
            success, message = database_manager.add_member_to_db(name, phone)
            # Use the returned success and message directly
            return success, message
        except Exception as e:
            # This catch block might be redundant if database_manager handles all its specific errors
            # and returns (False, error_message). However, it can catch unexpected errors.
            return False, f"An unexpected error occurred: {str(e)}"

    def save_plan_action(self, plan_name: str, duration_str: str, plan_id_to_update: str) -> tuple[bool, str, list | None]:
        """Saves a new plan or updates an existing one."""
        if not plan_name or not duration_str:
            return False, "Error: Plan Name and Duration cannot be empty.", None

        try:
            duration_days = int(duration_str)
            if duration_days <= 0:
                return False, "Error: Duration must be a positive integer.", None
        except ValueError:
            return False, "Error: Duration must be a valid integer.", None

        success = False
        message = ""
        updated_plans = None

        try:
            if plan_id_to_update:  # Editing existing plan
                plan_id = int(plan_id_to_update)
                # database_manager.update_plan now returns -> Tuple[bool, str]
                db_success, db_message = database_manager.update_plan(plan_id, plan_name, duration_days)
                success = db_success
                message = db_message
            else:  # Adding new plan
                # database_manager.add_plan now returns -> Tuple[bool, str, Optional[int]]
                db_success, db_message, returned_plan_id = database_manager.add_plan(plan_name, duration_days)
                success = db_success
                message = db_message
                # returned_plan_id is available if needed, e.g., for logging or immediate use,
                # but current logic just relies on success status.

            if success:
                updated_plans = database_manager.get_all_plans_with_inactive()
            return success, message, updated_plans
        except Exception as e:
            # Catch unexpected errors during the controller logic (e.g., int conversion if not validated)
            return False, f"An unexpected error occurred in save_plan_action: {str(e)}", None

    def toggle_plan_status_action(self, plan_id: int, current_status: bool) -> tuple[bool, str, list | None]:
        """Activates or deactivates a plan."""
        new_status = not current_status
        # database_manager.set_plan_active_status now returns -> Tuple[bool, str]
        db_success, db_message = database_manager.set_plan_active_status(plan_id, new_status)
        success = db_success
        message = db_message # Use the message from the database manager
        updated_plans = None

        if success:
            # Optionally, you could still use a custom success message or augment db_message
            # For now, directly using db_message. If a custom success message is preferred:
            # message = f"Plan status changed to {'Active' if new_status else 'Inactive'}."
            updated_plans = database_manager.get_all_plans_with_inactive()
        # If not successful, db_message already contains the error.
        return success, message, updated_plans

    def save_membership_action(self, membership_type: str, member_id: int | None,
                               start_date_str: str, amount_paid_str: str,
                               selected_plan_id: int | None = None,
                               payment_date_str: str | None = None,
                               payment_method: str | None = None,
                               sessions_str: str | None = None) -> tuple[bool, str]:
        """Handles the save membership action with validation based on membership type."""
        from datetime import datetime # Moved import

        # --- Basic Validation ---
        if not member_id: # Simplified check, assuming member_id is directly passed
            return False, "Error: Please select a valid member."

        if not start_date_str:
            return False, "Error: Start Date cannot be empty."
        try:
            datetime.strptime(start_date_str, '%Y-%m-%d')
        except ValueError:
            return False, "Error: Invalid Start Date format. Use YYYY-MM-DD."

        try:
            amount_paid = float(amount_paid_str)
            if amount_paid <= 0: # Changed from < to <=
                return False, "Error: Amount Paid must be a positive number."
        except ValueError:
            return False, "Error: Invalid Amount Paid. Must be a number."

        success = False
        try:
            if membership_type == "Group Class":
                if not selected_plan_id: # Simplified check
                    return False, "Error: Please select a valid plan."
                if not payment_date_str:
                    return False, "Error: Payment Date cannot be empty for Group Class."
                try:
                    datetime.strptime(payment_date_str, '%Y-%m-%d')
                except ValueError:
                    return False, "Error: Invalid Payment Date format. Use YYYY-MM-DD."
                if not payment_method: # Ensure payment_method is not None or empty
                    return False, "Error: Payment Method cannot be empty for Group Class."

                success, db_message = database_manager.add_transaction( # Expecting (bool, str)
                    transaction_type="Group Class",
                    member_id=member_id,
                    plan_id=selected_plan_id,
                    payment_date=payment_date_str,
                    start_date=start_date_str,
                    amount_paid=amount_paid,
                    payment_method=payment_method
                )
                if success:
                    return True, f"{membership_type} membership added successfully!"
                else:
                    return False, db_message

            elif membership_type == "Personal Training":
                if not sessions_str:
                    return False, "Error: Number of Sessions cannot be empty for PT."
                try:
                    sessions = int(sessions_str)
                    if sessions <= 0:
                        return False, "Error: Number of Sessions must be a positive integer."
                except ValueError:
                    return False, "Error: Number of Sessions must be an integer."

                success, db_message = database_manager.add_transaction( # Expecting (bool, str)
                    transaction_type="Personal Training",
                    member_id=member_id,
                    plan_id=None, # No plan_id for PT
                    payment_date=start_date_str, # For PT, payment_date is the start_date
                    start_date=start_date_str,
                    amount_paid=amount_paid,
                    payment_method="N/A", # Default or could be an argument if needed
                    sessions=sessions
                )
                if success:
                    return True, f"{membership_type} membership added successfully!"
                else:
                    return False, db_message
            else:
                return False, "Error: Unknown membership type selected."

        except Exception as e:
            # Log the exception e for debugging purposes if possible
            return False, f"An error occurred: {str(e)}"

    def generate_custom_pending_renewals_action(self, year: int, month: int) -> tuple[bool, str, list | None]:
        """Fetches pending renewals for a specific year and month."""
        import calendar # For month name in message

        try:
            renewals = database_manager.get_pending_renewals(year, month)
            month_name = calendar.month_name[month]

            if renewals:
                return True, f"Found {len(renewals)} pending renewals for {month_name} {year}.", renewals
            else:
                return True, f"No pending renewals found for {month_name} {year}.", []
        except Exception as e:
            month_name_str = str(month)
            try:
                month_name_str = calendar.month_name[month]
            except IndexError:
                pass
            return False, f"Error generating renewals report for {month_name_str} {year}: {str(e)}", None

    def generate_pending_renewals_action(self) -> tuple[bool, str, list | None]:
        """Fetches pending renewals for the current month."""
        from datetime import date
        import calendar

        today = date.today()
        current_year = today.year
        current_month = today.month
        month_name = calendar.month_name[current_month]

        try:
            renewals = database_manager.get_pending_renewals(year=current_year, month=current_month)
            if renewals:
                return True, f"Found {len(renewals)} pending renewals for {month_name} {current_year}.", renewals
            else:
                return True, f"No pending renewals found for {month_name} {current_year}.", []
        except Exception as e:
            return False, f"Error generating report for current month: {str(e)}", None

    def generate_finance_report_excel_action(self, year: int, month: int, save_path: str) -> tuple[bool, str]:
        """
        Generates an Excel finance report for the given month and year.
        The report includes a summary sheet and a detailed transactions sheet.
        Applies styling to the report.
        """
        try:
            transactions_data = database_manager.get_transactions_for_month(year, month)
            month_name_str = calendar.month_name[month]

            if not transactions_data:
                return True, f"No transaction data found for {month_name_str} {year}. Report not generated."

            df_columns = [
                "Transaction ID", "Client Name", "Payment Date", "Start Date", "End Date",
                "Amount Paid", "Type", "Plan/Sessions", "Payment Method"
            ]
            df = pd.DataFrame(transactions_data, columns=df_columns)

            df["Amount Paid"] = pd.to_numeric(df["Amount Paid"], errors='coerce').fillna(0)
            df["Payment Date"] = pd.to_datetime(df["Payment Date"], errors='coerce').dt.strftime('%Y-%m-%d')
            df["Start Date"] = pd.to_datetime(df["Start Date"], errors='coerce').dt.strftime('%Y-%m-%d')

            total_revenue = df["Amount Paid"].sum()
            total_transactions = len(df)

            summary_data = {
                "Metric": ["Total Revenue", "Total Transactions"],
                "Value": [total_revenue, total_transactions]
            }
            summary_df = pd.DataFrame(summary_data)

            with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
                summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=2)
                df.to_excel(writer, sheet_name="Detailed Transactions", index=False, startrow=1)

                workbook = writer.book
                summary_sheet = writer.sheets["Summary"]
                details_sheet = writer.sheets["Detailed Transactions"]

                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                title_font = Font(bold=True, size=16)
                border_style = Side(style="thin", color="000000")
                thin_border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

                summary_sheet.cell(row=1, column=1, value=f"Finance Summary - {month_name_str} {year}").font = title_font
                for col_num, column_title in enumerate(summary_df.columns, 1):
                    cell = summary_sheet.cell(row=3, column=col_num)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                for row_idx_offset, _ in enumerate(summary_df.index):
                    for col_idx, _ in enumerate(summary_df.columns):
                        summary_sheet.cell(row=row_idx_offset + 4, column=col_idx + 1).border = thin_border

                summary_sheet.column_dimensions['A'].width = 25
                summary_sheet.column_dimensions['B'].width = 15
                summary_sheet.cell(row=4, column=2).number_format = '"$"#,##0.00'

                details_sheet.cell(row=1, column=1, value=f"Detailed Transactions - {month_name_str} {year}").font = title_font
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = details_sheet.cell(row=2, column=col_num)
                    cell.value = column_title
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    column_letter = get_column_letter(col_num)
                    try:
                        max_len = max(df[column_title].astype(str).map(len).max(), len(str(column_title))) + 2
                    except (TypeError, ValueError):
                        max_len = len(str(column_title)) + 2
                    details_sheet.column_dimensions[column_letter].width = max_len if max_len < 50 else 50

                for row_idx_offset, _ in enumerate(df.index):
                    for col_idx, _ in enumerate(df.columns):
                        details_sheet.cell(row=row_idx_offset + 3, column=col_idx + 1).border = thin_border

                amount_paid_col_letter = get_column_letter(df_columns.index("Amount Paid") + 1)
                for row_num in range(3, details_sheet.max_row + 1):
                     details_sheet[f'{amount_paid_col_letter}{row_num}'].number_format = '"$"#,##0.00'

            return True, f"Finance report generated successfully: {save_path}"

        except ImportError:
            return False, "Error: pandas or openpyxl library not found. Please ensure they are installed."
        except Exception as e:
            return False, f"An error occurred during report generation: {str(e)}"

    def get_book_status_action(self, year: int, month: int) -> str:
        """
        Constructs month_key, calls database_manager.get_book_status, and returns a user-friendly message.
        """
        try:
            month_key = f"{year:04d}-{month:02d}"
            status = database_manager.get_book_status(month_key)
            return f"Status for {month_key}: {status.upper()}"
        except Exception as e:
            return f"Error getting book status: {str(e)}"

    def close_books_action(self, year: int, month: int) -> tuple[bool, str]:
        """
        Constructs month_key, calls database_manager.set_book_status to "closed", and returns status.
        """
        try:
            month_key = f"{year:04d}-{month:02d}"
            success = database_manager.set_book_status(month_key, "closed")
            if success:
                return True, f"Books for {month_key} closed successfully."
            else:
                return False, f"Failed to close books for {month_key}."
        except Exception as e:
            return False, f"Error closing books for {month_key}: {str(e)}"

    def open_books_action(self, year: int, month: int) -> tuple[bool, str]:
        """
        Constructs month_key, calls database_manager.set_book_status to "open", and returns status.
        """
        try:
            month_key = f"{year:04d}-{month:02d}"
            success = database_manager.set_book_status(month_key, "open")
            if success:
                return True, f"Books for {month_key} re-opened successfully."
            else:
                return False, f"Failed to re-open books for {month_key}."
        except Exception as e:
            return False, f"Error re-opening books for {month_key}: {str(e)}"

    def get_filtered_members(self, name_filter: Optional[str], phone_filter: Optional[str]) -> list:
        """Fetches members based on name and/or phone filters."""
        return database_manager.get_all_members(name_filter=name_filter, phone_filter=phone_filter)

    def get_filtered_transaction_history(self, name_filter: Optional[str], phone_filter: Optional[str], join_date_filter: Optional[str]) -> list:
        """Fetches transaction history based on name, phone, and/or join date filters."""
        return database_manager.get_transactions_with_member_details(
            name_filter=name_filter,
            phone_filter=phone_filter,
            join_date_filter=join_date_filter
        )

    def get_active_plans(self) -> list:
        """Fetches all active plans."""
        return database_manager.get_all_plans()

    def get_all_plans_with_inactive(self) -> list:
        """Fetches all plans, including inactive ones."""
        return database_manager.get_all_plans_with_inactive()

    def get_all_activity_for_member(self, member_id: int) -> list:
        """Fetches all activity records for a given member_id."""
        return database_manager.get_all_activity_for_member(member_id)

    def deactivate_member_action(self, member_id: int) -> tuple[bool, str]:
        """Handles the action of deactivating a member."""
        try:
            success, message = database_manager.deactivate_member(member_id)
            return success, message
        except Exception as e:
            return False, f"An unexpected error occurred while deactivating the member: {str(e)}"

    def delete_transaction_action(self, transaction_id: int) -> tuple[bool, str]:
        """Calls database_manager.delete_transaction and returns status."""
        try:
            success, message = database_manager.delete_transaction(transaction_id)
            if success: # delete_transaction returns (bool, str)
                return True, "Transaction deleted successfully." # Standardize success message
            else:
                return False, message # Pass through the error message
        except Exception as e:
            return False, f"An error occurred while deleting the transaction: {str(e)}"

    def delete_plan_action(self, plan_id: int) -> tuple[bool, str]:
        """Calls database_manager.delete_plan and returns its result or an error status."""
        try:
            success, message = database_manager.delete_plan(plan_id)
            return success, message
        except Exception as e:
            return False, f"An error occurred while deleting the plan: {str(e)}"


from .components.membership_tab import MembershipTab
from .components.history_tab import HistoryTab
from .components.plans_tab import PlansTab
from .components.reporting_tab import ReportingTab
from .components.settings_tab import SettingsTab

class FletAppView(ft.Container):
    def __init__(self):
        super().__init__() # Calls ft.Container.__init__
        self.controller = GuiController()
        # self.selected_member_id_flet: Optional[int] = None # Moved to MembershipTab
        self.selected_transaction_id_flet: Optional[int] = None
        self.selected_plan_id_flet: Optional[int] = None
        self.current_plan_id_to_update_flet: Optional[int] = None
        # self.selected_start_date_flet: Optional[date] = None # Managed by MembershipTab locally
        # self.selected_payment_date_flet: Optional[date] = None # Managed by MembershipTab locally
        # self.active_date_picker_target: Optional[str] = None # Managed by MembershipTab locally
        # self.page is available via self.page once the control is added to a page.

        # UI declarations
        # self.member_actions_feedback_text: Optional[ft.Text] = None # Moved to MembershipTab
        self.history_actions_feedback_text: Optional[ft.Text] = None
        # self.delete_member_button_flet: Optional[ft.ElevatedButton] = None # Moved to MembershipTab
        self.delete_plan_button_flet: Optional[ft.ElevatedButton] = None
        self.delete_transaction_button_flet: Optional[ft.ElevatedButton] = None
        # self.member_name_input: Optional[ft.TextField] = None # Moved to MembershipTab
        # self.member_phone_input: Optional[ft.TextField] = None # Moved to MembershipTab
        # self.add_member_button: Optional[ft.ElevatedButton] = None # Moved to MembershipTab
        # self.member_form_feedback_text: Optional[ft.Text] = None # Moved to MembershipTab (specific instance)
        self.plan_name_input: Optional[ft.TextField] = None
        self.plan_duration_input: Optional[ft.TextField] = None
        self.save_plan_button: Optional[ft.ElevatedButton] = None
        self.plan_form_feedback_text: Optional[ft.Text] = None
        self.edit_plan_button: Optional[ft.ElevatedButton] = None
        self.clear_plan_form_button: Optional[ft.ElevatedButton] = None
        self.toggle_plan_status_button: Optional[ft.ElevatedButton] = None
        self.history_name_filter_input: Optional[ft.TextField] = None
        self.history_phone_filter_input: Optional[ft.TextField] = None
        self.history_join_date_filter_input: Optional[ft.TextField] = None
        self.apply_history_filters_button: Optional[ft.ElevatedButton] = None
        self.clear_history_filters_button: Optional[ft.ElevatedButton] = None
        # self.membership_type_dropdown: Optional[ft.Dropdown] = None # Moved to MembershipTab
        # self.membership_member_dropdown: Optional[ft.Dropdown] = None # Moved to MembershipTab
        # self.membership_plan_dropdown: Optional[ft.Dropdown] = None # Moved to MembershipTab
        # self.membership_sessions_input: Optional[ft.TextField] = None # Moved to MembershipTab
        # self.membership_start_date_picker_button: Optional[ft.ElevatedButton] = None # Moved to MembershipTab
        # self.membership_start_date_text: Optional[ft.Text] = None # Moved to MembershipTab
        # self.membership_payment_date_picker_button: Optional[ft.ElevatedButton] = None # Moved to MembershipTab
        # self.membership_payment_date_text: Optional[ft.Text] = None # Moved to MembershipTab
        # self.membership_amount_paid_input: Optional[ft.TextField] = None # Moved to MembershipTab
        # self.membership_payment_method_input: Optional[ft.TextField] = None # Moved to MembershipTab
        # self.save_membership_button: Optional[ft.ElevatedButton] = None # Moved to MembershipTab
        # self.membership_form_feedback_text: Optional[ft.Text] = None # Moved to MembershipTab (specific instance)
        self.renewal_report_year_input: Optional[ft.TextField] = None
        self.renewal_report_month_dropdown: Optional[ft.Dropdown] = None
        self.generate_renewals_report_button: Optional[ft.ElevatedButton] = None
        self.renewals_report_feedback_text: Optional[ft.Text] = None
        self.finance_report_year_input: Optional[ft.TextField] = None
        self.finance_report_month_dropdown: Optional[ft.Dropdown] = None
        self.generate_finance_report_button: Optional[ft.ElevatedButton] = None
        self.finance_report_feedback_text: Optional[ft.Text] = None
        # self.members_table_flet: Optional[ft.DataTable] = None # Moved to MembershipTab
        # self.member_specific_history_table_flet: Optional[ft.DataTable] = None # Moved to MembershipTab
        self.full_history_table_flet: Optional[ft.DataTable] = None
        self.plans_table_flet: Optional[ft.DataTable] = None
        self.pending_renewals_table_flet: Optional[ft.DataTable] = None
        self.tabs_control: Optional[ft.Tabs] = None

        # Date Picker and File Picker are declared here and initialized in build
        self.date_picker: Optional[ft.DatePicker] = None
        self.file_picker: Optional[ft.FilePicker] = None
        self.expand = True # FletAppView itself should expand

        # Date Picker and File Picker are initialized here, they are shared across tabs if needed
        # Handlers for DatePicker are now managed by individual tabs if they use it
        self.date_picker = ft.DatePicker()
        self.file_picker = ft.FilePicker(on_result=self.on_file_picker_result_flet)


    # Event Handlers and Helper Methods (methods specific to MembershipTab have been moved)
    def on_full_history_select_changed(self, e: ft.ControlEvent):
        """Handles row selection changes in the full_history_table_flet."""
        selected_index_str = e.data
        if selected_index_str:
            try:
                selected_index = int(selected_index_str)
                if 0 <= selected_index < len(self.full_history_table_flet.rows):
                    selected_row = self.full_history_table_flet.rows[selected_index]
                    txn_id_cell = selected_row.cells[0]
                    if isinstance(txn_id_cell.content, ft.Text):
                        self.selected_transaction_id_flet = int(txn_id_cell.content.value)
                    else:
                        self.selected_transaction_id_flet = None
                else:
                    self.selected_transaction_id_flet = None
            except ValueError:
                self.selected_transaction_id_flet = None
            except Exception:
                self.selected_transaction_id_flet = None
        else:
            self.selected_transaction_id_flet = None
        # print(f"DEBUG: FletAppView - Selected Transaction ID: {self.selected_transaction_id_flet}")
        # self.update()

    def on_plan_select_changed(self, e: ft.ControlEvent):
        """Handles row selection changes in the plans_table_flet."""
        selected_index_str = e.data
        if selected_index_str:
            try:
                selected_index = int(selected_index_str)
                if 0 <= selected_index < len(self.plans_table_flet.rows):
                    selected_row = self.plans_table_flet.rows[selected_index]
                    plan_id_cell = selected_row.cells[0]
                    if isinstance(plan_id_cell.content, ft.Text):
                        self.selected_plan_id_flet = int(plan_id_cell.content.value)
                    else:
                        self.selected_plan_id_flet = None
                else:
                    self.selected_plan_id_flet = None
            except ValueError:
                self.selected_plan_id_flet = None
            except Exception:
                self.selected_plan_id_flet = None
        else:
            self.selected_plan_id_flet = None

        if self.selected_plan_id_flet is not None:
            self.toggle_plan_status_button.disabled = False
        else:
            self.toggle_plan_status_button.disabled = True

        if hasattr(self, 'toggle_plan_status_button') and self.toggle_plan_status_button.page:
            self.toggle_plan_status_button.update()
        # print(f"DEBUG: FletAppView - Selected Plan ID: {self.selected_plan_id_flet}")
        # self.update()

    def _get_membership_status_flet(self, end_date_str: Optional[str]) -> str:
        """Helper to determine membership status based on end date string.
           This can remain if other tabs use it, or be moved/duplicated if only used by one tab's logic.
           For now, keeping it here as full_history_table_flet uses it.
        """
        if not end_date_str or end_date_str.lower() == "n/a" or end_date_str.strip() == "":
            return "N/A"
        try:
            end_date_obj = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            return "Active" if end_date_obj >= date.today() else "Inactive"
        except ValueError:
            return "Invalid Date"

    def refresh_membership_history_display_flet(self, transactions_list: Optional[list] = None):
        """Populates the full_history_table_flet with transaction data."""
        if not hasattr(self, 'full_history_table_flet') or self.full_history_table_flet is None:
            return # Table not initialized yet

        self.full_history_table_flet.rows.clear()
        if transactions_list is None:
            transactions_data = self.controller.get_filtered_transaction_history(None, None, None)
        else:
            transactions_data = transactions_list

        if not transactions_data:
            self.full_history_table_flet.rows.append(
                ft.DataRow(cells=[ft.DataCell(ft.Text("No transaction records found."), colspan=len(self.full_history_table_flet.columns))])
            )
        else:
            for record in transactions_data:
                (transaction_id, _member_id, transaction_type, plan_id, payment_date,
                 start_date, end_date, amount_paid, payment_method_db, sessions,
                 client_name, phone, join_date) = record
                plan_id_or_sessions_display = str(plan_id) if transaction_type == "Group Class" and plan_id is not None else (str(sessions) if transaction_type == "Personal Training" and sessions is not None else "N/A")
                amount_paid_formatted = f"{float(amount_paid):.2f}" if amount_paid is not None else "0.00"
                end_date_display = str(end_date) if end_date is not None else "N/A"
                status = self._get_membership_status_flet(str(end_date) if end_date else None)
                ordered_values = [
                    str(transaction_id), str(client_name or "N/A"), str(phone or "N/A"), str(join_date or "N/A"),
                    str(transaction_type or "N/A"), amount_paid_formatted, str(payment_date or "N/A"),
                    str(start_date or "N/A"), end_date_display, status, plan_id_or_sessions_display,
                    str(payment_method_db or "N/A")
                ]
                data_cells = [ft.DataCell(ft.Text(value)) for value in ordered_values]
                self.full_history_table_flet.rows.append(ft.DataRow(cells=data_cells))

        if self.full_history_table_flet.page: self.full_history_table_flet.update()
        # self.update()

    def apply_history_filters_flet(self, e):
        name_filter = self.history_name_filter_input.value or None
        phone_filter = self.history_phone_filter_input.value or None
        join_date_filter = self.history_join_date_filter_input.value or None
        if join_date_filter:
            try:
                datetime.strptime(join_date_filter, '%Y-%m-%d')
            except ValueError:
                self.history_actions_feedback_text.value = "Error: Invalid Join Date format. Use YYYY-MM-DD."
                self.history_actions_feedback_text.color = ft.colors.RED
                if self.history_actions_feedback_text.page: self.history_actions_feedback_text.update()
                return
        filtered_data = self.controller.get_filtered_transaction_history(name_filter, phone_filter, join_date_filter)
        self.refresh_membership_history_display_flet(filtered_data)
        feedback_msg = f"Filters applied. Found {len(filtered_data)} records." if filtered_data else "No results for current filters."
        self.history_actions_feedback_text.value = feedback_msg
        self.history_actions_feedback_text.color = ft.colors.GREEN if filtered_data else ft.colors.ORANGE
        if self.history_actions_feedback_text.page: self.history_actions_feedback_text.update()
        # self.update()

    def clear_history_filters_flet(self, e):
        self.history_name_filter_input.value = ""
        self.history_phone_filter_input.value = ""
        self.history_join_date_filter_input.value = ""
        if self.history_name_filter_input.page: self.history_name_filter_input.update()
        if self.history_phone_filter_input.page: self.history_phone_filter_input.update()
        if self.history_join_date_filter_input.page: self.history_join_date_filter_input.update()
        self.refresh_membership_history_display_flet()
        self.history_actions_feedback_text.value = "Filters cleared. Displaying all history."
        self.history_actions_feedback_text.color = ft.colors.BLUE
        if self.history_actions_feedback_text.page: self.history_actions_feedback_text.update()
        # self.update()

    def display_all_plans_flet(self, plans_list: Optional[list] = None):
        if not hasattr(self, 'plans_table_flet') or self.plans_table_flet is None: return
        if plans_list is None:
            plans_list = self.controller.get_all_plans_with_inactive()
        self.plans_table_flet.rows.clear()
        if plans_list:
            for plan_data in plans_list:
                plan_id, plan_name, duration_days, is_active = plan_data
                status = "Active" if is_active else "Inactive"
                row = ft.DataRow(cells=[
                    ft.DataCell(ft.Text(str(plan_id))), ft.DataCell(ft.Text(str(plan_name))),
                    ft.DataCell(ft.Text(str(duration_days))), ft.DataCell(ft.Text(status)),
                ])
                self.plans_table_flet.rows.append(row)
        else:
            self.plans_table_flet.rows.append(
                ft.DataRow(cells=[ft.DataCell(ft.Text("No plans found."), colspan=len(self.plans_table_flet.columns))])
            )
        if self.plans_table_flet.page: self.plans_table_flet.update()
        # self.update()

    def display_pending_renewals_flet(self): # This seems to be for the "Reporting" tab, keep here.
        if not hasattr(self, 'pending_renewals_table_flet') or self.pending_renewals_table_flet is None: return
        success, message, renewals_data = self.controller.generate_pending_renewals_action()
        self.pending_renewals_table_flet.rows.clear()
        if success and renewals_data:
            for renewal_item in renewals_data:
                cells = [ft.DataCell(ft.Text(str(item))) for item in renewal_item]
                self.pending_renewals_table_flet.rows.append(ft.DataRow(cells=cells))
        else:
            display_message = message if not success else "No pending renewals found for the current month."
            self.pending_renewals_table_flet.rows.append(
                ft.DataRow(cells=[ft.DataCell(ft.Text(display_message), colspan=len(self.pending_renewals_table_flet.columns))])
            )
        if self.pending_renewals_table_flet.page: self.pending_renewals_table_flet.update()
        # self.update()

    def on_save_plan_click(self, e):
        plan_name = self.plan_name_input.value
        duration_str = self.plan_duration_input.value
        plan_id_to_update_str = str(self.current_plan_id_to_update_flet) if self.current_plan_id_to_update_flet else None
        success, message, updated_plans = self.controller.save_plan_action(plan_name, duration_str, plan_id_to_update_str)
        self.plan_form_feedback_text.value = message
        if success:
            self.plan_form_feedback_text.color = ft.colors.GREEN
            self.on_clear_plan_form_click(None)
            if updated_plans is not None: self.display_all_plans_flet(updated_plans)
            else: self.display_all_plans_flet()
            # self.populate_plan_dropdowns_flet() # This method is now in MembershipTab
            # If MembershipTab is already instantiated and needs update, it should handle its own dropdown refresh.
            # Or FletAppView needs a reference to call it. For now, assume MembershipTab handles this.
            if hasattr(self, 'membership_tab_ref') and self.membership_tab_ref:
                 self.membership_tab_ref.populate_plan_dropdowns_flet()

        else:
            self.plan_form_feedback_text.color = ft.colors.RED
        if self.plan_form_feedback_text.page: self.plan_form_feedback_text.update()
        # self.update()

    def on_edit_selected_plan_click(self, e):
        if self.selected_plan_id_flet is not None:
            selected_plan_details = None
            for row in self.plans_table_flet.rows:
                if int(row.cells[0].content.value) == self.selected_plan_id_flet:
                    selected_plan_details = {"id": row.cells[0].content.value, "name": row.cells[1].content.value, "duration": row.cells[2].content.value}
                    break
            if selected_plan_details:
                self.plan_name_input.value = selected_plan_details["name"]
                self.plan_duration_input.value = selected_plan_details["duration"]
                self.current_plan_id_to_update_flet = self.selected_plan_id_flet
                self.plan_form_feedback_text.value = f"Editing Plan ID: {self.current_plan_id_to_update_flet}"
                self.plan_form_feedback_text.color = ft.colors.BLUE
            else:
                self.plan_form_feedback_text.value = "Error: Could not find details for selected plan."
                self.plan_form_feedback_text.color = ft.colors.RED
        else:
            self.plan_form_feedback_text.value = "Please select a plan to edit."
            self.plan_form_feedback_text.color = ft.colors.ORANGE
        if self.plan_name_input.page: self.plan_name_input.update()
        if self.plan_duration_input.page: self.plan_duration_input.update()
        if self.plan_form_feedback_text.page: self.plan_form_feedback_text.update()
        # self.update()

    def on_clear_plan_form_click(self, e):
        self.plan_name_input.value = ""
        self.plan_duration_input.value = ""
        self.current_plan_id_to_update_flet = None
        self.plan_form_feedback_text.value = "Ready to add a new plan."
        self.plan_form_feedback_text.color = ft.colors.BLACK
        if self.plan_name_input.page: self.plan_name_input.update()
        if self.plan_duration_input.page: self.plan_duration_input.update()
        if self.plan_form_feedback_text.page: self.plan_form_feedback_text.update()
        # self.update()

    # populate_plan_dropdowns_flet removed (moved to MembershipTab)
    # open_date_picker removed (functionality moved to MembershipTab's local date picker)
    # on_date_picker_change removed
    # on_date_picker_dismiss removed
    # on_membership_type_change_flet removed (moved to MembershipTab)
    # on_save_membership_click removed (moved to MembershipTab)

    def on_generate_renewals_report_click(self, e): # For Reporting Tab
        year_str = self.renewal_report_year_input.value
        month_str = self.renewal_report_month_dropdown.value
        if not year_str or not month_str:
            self.renewals_report_feedback_text.value = "Year and Month cannot be empty."
            self.renewals_report_feedback_text.color = ft.colors.RED
            if self.renewals_report_feedback_text.page: self.renewals_report_feedback_text.update()
            return
        try:
            year = int(year_str); month = int(month_str)
            if not (1 <= month <= 12): raise ValueError("Month out of range.")
        except ValueError:
            self.renewals_report_feedback_text.value = "Invalid Year or Month format."
            self.renewals_report_feedback_text.color = ft.colors.RED
            if self.renewals_report_feedback_text.page: self.renewals_report_feedback_text.update()
            return
        success, message, renewals_data = self.controller.generate_custom_pending_renewals_action(year, month)
        self.renewals_report_feedback_text.value = message
        self.renewals_report_feedback_text.color = ft.colors.GREEN if success else ft.colors.RED
        if self.renewals_report_feedback_text.page: self.renewals_report_feedback_text.update()
        self.pending_renewals_table_flet.rows.clear()
        if success and renewals_data:
            for item in renewals_data:
                cells = [ft.DataCell(ft.Text(str(val))) for val in item]
                self.pending_renewals_table_flet.rows.append(ft.DataRow(cells=cells))
        elif success and not renewals_data:
             self.pending_renewals_table_flet.rows.append(ft.DataRow(cells=[ft.DataCell(ft.Text(message), colspan=len(self.pending_renewals_table_flet.columns))]))
        else:
            error_message = message if message else "Error generating report or no data."
            self.pending_renewals_table_flet.rows.append(ft.DataRow(cells=[ft.DataCell(ft.Text(error_message), colspan=len(self.pending_renewals_table_flet.columns))]))
        if self.pending_renewals_table_flet.page: self.pending_renewals_table_flet.update()
        # self.update()

    def on_generate_finance_report_click(self, e): # For Reporting Tab
        year_str = self.finance_report_year_input.value
        month_str = self.finance_report_month_dropdown.value
        if not year_str or not month_str:
            self.finance_report_feedback_text.value = "Year and Month cannot be empty."
            self.finance_report_feedback_text.color = ft.colors.RED
            if self.finance_report_feedback_text.page: self.finance_report_feedback_text.update()
            return
        try:
            int(year_str);
            if not (1 <= int(month_str) <= 12): raise ValueError("Month out of range.")
        except ValueError:
            self.finance_report_feedback_text.value = "Invalid Year or Month format."
            self.finance_report_feedback_text.color = ft.colors.RED
            if self.finance_report_feedback_text.page: self.finance_report_feedback_text.update()
            return
        self.file_picker.save_file(
            dialog_title="Save Finance Report",
            file_name=f"finance_report_{year_str}_{month_str}.xlsx",
            allowed_extensions=["xlsx"]
        )

    def on_file_picker_result_flet(self, e: ft.FilePickerResultEvent): # Used by Finance Report
        if e.path:
            save_path = e.path
            year = int(self.finance_report_year_input.value)
            month = int(self.finance_report_month_dropdown.value)
            success, message = self.controller.generate_finance_report_excel_action(year, month, save_path)
            self.finance_report_feedback_text.value = message
            self.finance_report_feedback_text.color = ft.colors.GREEN if success else ft.colors.RED
        else:
            self.finance_report_feedback_text.value = "File save cancelled."
            self.finance_report_feedback_text.color = ft.colors.ORANGE
        if self.finance_report_feedback_text.page: self.finance_report_feedback_text.update()
        # self.update()

    def build(self):
        # DatePicker and FilePicker are already initialized in __init__
        # self.date_picker = ft.DatePicker(...) # Removed specific handlers
        # self.file_picker = ft.FilePicker(on_result=self.on_file_picker_result_flet) # Already in __init__

        # UI Element Initializations for controls remaining in FletAppView
        # self.member_actions_feedback_text = ft.Text("") # Moved
        self.history_actions_feedback_text = ft.Text("")

        # self.delete_member_button_flet = ft.ElevatedButton(...) # Moved
        self.delete_plan_button_flet = ft.ElevatedButton(text="Delete Selected Plan", on_click=self.on_delete_selected_plan_click_flet)
        self.delete_transaction_button_flet = ft.ElevatedButton(text="Delete Selected Transaction", on_click=self.on_delete_selected_transaction_click_flet)

        # Member form controls moved to MembershipTab
        # Plan form controls remain for Plan Management Tab
        self.plan_name_input = ft.TextField(label="Plan Name")
        self.plan_duration_input = ft.TextField(label="Duration (days)")
        self.save_plan_button = ft.ElevatedButton(text="Save Plan", on_click=self.on_save_plan_click)
        self.plan_form_feedback_text = ft.Text("Ready to add a new plan.", color=ft.colors.BLACK)
        self.edit_plan_button = ft.ElevatedButton(text="Edit Selected Plan", on_click=self.on_edit_selected_plan_click)
        self.clear_plan_form_button = ft.ElevatedButton(text="Clear Form / New Plan", on_click=self.on_clear_plan_form_click)
        self.toggle_plan_status_button = ft.ElevatedButton(text="Toggle Active/Inactive", on_click=self.on_toggle_plan_status_click, disabled=True)

        # History filter controls remain for Membership History Tab
        self.history_name_filter_input = ft.TextField(label="Filter by Name")
        self.history_phone_filter_input = ft.TextField(label="Filter by Phone")
        self.history_join_date_filter_input = ft.TextField(label="Filter by Join Date (YYYY-MM-DD)")
        self.apply_history_filters_button = ft.ElevatedButton(text="Apply History Filters", on_click=self.apply_history_filters_flet)
        self.clear_history_filters_button = ft.ElevatedButton(text="Clear History Filters", on_click=self.clear_history_filters_flet)

        # Membership form controls moved to MembershipTab

        # Reporting controls remain for Reporting Tab
        self.renewal_report_year_input = ft.TextField(label="Year", value=str(datetime.now().year))
        self.renewal_report_month_dropdown = ft.Dropdown(
            label="Month", options=[ft.dropdown.Option(str(i), str(i)) for i in range(1, 13)], value=str(datetime.now().month)
        )
        self.generate_renewals_report_button = ft.ElevatedButton(text="Generate Renewals Report", on_click=self.on_generate_renewals_report_click)
        self.renewals_report_feedback_text = ft.Text("")

        self.finance_report_year_input = ft.TextField(label="Year", value=str(datetime.now().year))
        self.finance_report_month_dropdown = ft.Dropdown(
            label="Month", options=[ft.dropdown.Option(str(i), str(i)) for i in range(1, 13)], value=str(datetime.now().month)
        )
        self.generate_finance_report_button = ft.ElevatedButton(text="Generate Excel Finance Report", on_click=self.on_generate_finance_report_click)
        self.finance_report_feedback_text = ft.Text("")

        # Tables: members_table_flet and member_specific_history_table_flet moved to MembershipTab
        # Full history table remains for Membership History Tab
        self.full_history_table_flet = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("TXN ID")), ft.DataColumn(ft.Text("Name")), ft.DataColumn(ft.Text("Phone")),
                ft.DataColumn(ft.Text("Joined")), ft.DataColumn(ft.Text("Type")), ft.DataColumn(ft.Text("Amount ($)")),
                ft.DataColumn(ft.Text("Paid Date")), ft.DataColumn(ft.Text("Start Date")), ft.DataColumn(ft.Text("End Date")),
                ft.DataColumn(ft.Text("Status")), ft.DataColumn(ft.Text("Plan/Sessions")), ft.DataColumn(ft.Text("Pay Method")),
            ],
            rows=[], on_select_changed=self.on_full_history_select_changed
        )
        # Plans table remains for Plan Management Tab
        self.plans_table_flet = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("ID")), ft.DataColumn(ft.Text("Name")),
                ft.DataColumn(ft.Text("Duration (Days)")), ft.DataColumn(ft.Text("Status")),
            ],
            rows=[], on_select_changed=self.on_plan_select_changed
        )
        # Pending renewals table remains for Reporting Tab
        self.pending_renewals_table_flet = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Member ID")), ft.DataColumn(ft.Text("Name")), ft.DataColumn(ft.Text("Phone")),
                ft.DataColumn(ft.Text("Plan Name")), ft.DataColumn(ft.Text("End Date")), ft.DataColumn(ft.Text("Days Overdue")),
            ],
            rows=[]
        )

        # Instantiate MembershipTab
        self.membership_tab_ref = MembershipTab(self.controller, self.date_picker)
        # Instantiate SettingsTab
        self.settings_tab_ref = SettingsTab(self.controller)


        # UI Assembly
        # table_area_container related to members_table_flet is now part of MembershipTab.build()

        self.tabs_control = ft.Tabs(
            tabs=[
                ft.Tab(
                    text="Membership Management",
                    content=self.membership_tab_ref # Use the instance of MembershipTab
                ),
                ft.Tab(
                    text="Membership History",
                    content=ft.Column(
                        controls=[
                            ft.Container(
                                content=ft.Column(
                                    controls=[
                                        ft.Text("Filter Transaction History", weight=ft.FontWeight.BOLD),
                                        self.history_name_filter_input,
                                        self.history_phone_filter_input,
                                        self.history_join_date_filter_input,
                                        ft.Row(
                                            controls=[self.apply_history_filters_button, self.clear_history_filters_button],
                                            alignment=ft.MainAxisAlignment.START,
                                        ),
                                    ], spacing=10
                                ), bgcolor=ft.colors.GREEN_200, padding=10, border_radius=5,
                            ),
                            ft.Container(
                                content=ft.Column(controls=[self.full_history_table_flet], scroll=ft.ScrollMode.AUTO, expand=True),
                                expand=4, bgcolor=ft.colors.GREEN_300, padding=10, alignment=ft.alignment.top_center
                            ),
                            self.delete_transaction_button_flet,
                            self.history_actions_feedback_text,
                        ], spacing=10
                    )
                ),
                ft.Tab(
                    text="Plan Management",
                    content=ft.Row(
                        controls=[
                            ft.Container(
                                content=ft.Column(
                                    controls=[
                                        ft.Text("Add/Edit Plan", weight=ft.FontWeight.BOLD),
                                        self.plan_name_input,
                                        self.plan_duration_input,
                                        self.save_plan_button,
                                        self.edit_plan_button,
                                        self.clear_plan_form_button,
                                        self.toggle_plan_status_button,
                                        self.delete_plan_button_flet,
                                        self.plan_form_feedback_text,
                                    ]
                                ), expand=1, bgcolor=ft.colors.AMBER_200, padding=10, alignment=ft.alignment.top_center,
                            ),
                            ft.Container(
                                content=ft.Column(controls=[self.plans_table_flet], scroll=ft.ScrollMode.AUTO, expand=True),
                                expand=2, bgcolor=ft.colors.AMBER_300, padding=10, alignment=ft.alignment.top_center
                            ),
                        ]
                    )
                ),
                ft.Tab(
                    text="Reporting",
                    content=ft.Column(
                        controls=[
                            ft.Container(
                                content=ft.Column(
                                    controls=[
                                        ft.Text("Custom Pending Renewals Report", weight=ft.FontWeight.BOLD),
                                        self.renewal_report_year_input,
                                        self.renewal_report_month_dropdown,
                                        self.generate_renewals_report_button,
                                        self.renewals_report_feedback_text,
                                        ft.Divider(),
                                        ft.Text("Pending Renewals Results", weight=ft.FontWeight.BOLD),
                                        self.pending_renewals_table_flet
                                    ], scroll=ft.ScrollMode.AUTO, expand=True
                                ), expand=1, bgcolor=ft.colors.TEAL_200, padding=10, alignment=ft.alignment.top_center
                            ),
                            ft.Container(
                                content=ft.Column(
                                    controls=[
                                        ft.Text("Generate Monthly Finance Report (Excel)", weight=ft.FontWeight.BOLD),
                                        self.finance_report_year_input,
                                        self.finance_report_month_dropdown,
                                        self.generate_finance_report_button,
                                        self.finance_report_feedback_text,
                                    ]
                                ), expand=1, bgcolor=ft.colors.TEAL_300, padding=10, alignment=ft.alignment.top_center,
                            ),
                        ]
                    )
                ),
                ft.Tab(
                    text="Settings",
                    content=self.settings_tab_ref
                ),
            ]
        )

        # Initial data population for tables remaining in FletAppView
        # self.display_all_members_flet() # Moved to MembershipTab
        self.refresh_membership_history_display_flet() # For full history table
        self.display_all_plans_flet() # For plans table

        # Populate dropdowns related to MembershipTab are handled within MembershipTab itself.
        # self.populate_member_dropdowns_flet() # Moved to MembershipTab
        # self.populate_plan_dropdowns_flet() # Moved to MembershipTab & also called from on_save_plan_click if plans change

        # Set initial visibility for membership form - handled by MembershipTab.build()
        # self.on_membership_type_change_flet(None) # Moved to MembershipTab

        # Add DatePicker to the page overlay. This is done in did_mount.
        # Initial population of the pending renewals table with default/current month's data
        self.on_generate_renewals_report_click(None) # For Reporting Tab

        return self.tabs_control

    def did_mount(self):
        """Called after the control is added to the page."""
        if self.page:
            if self.date_picker not in self.page.overlay: # Shared date_picker
                self.page.overlay.append(self.date_picker)
            if self.file_picker not in self.page.overlay: # Shared file_picker
                self.page.overlay.append(self.file_picker)
            self.page.update()

def main(page: ft.Page):
    page.title = "Kranos MMA Reporter"
    page.theme_mode = ft.ThemeMode.DARK

    # Create and add the main app view
    app_view = FletAppView()
    page.add(app_view) # For a container, this adds it. Content was set in __init__. Overlay logic in did_mount.

    # Overlay components are added in did_mount, so no need to do it here explicitly
    # if app_view.date_picker not in page.overlay:
    #     page.overlay.append(app_view.date_picker)
    # if app_view.file_picker not in page.overlay: # Ensure FilePicker is added
    #     page.overlay.append(app_view.file_picker)
    # page.update() # page.add will trigger an update

if __name__ == "__main__":
    # Ensure database is initialized
    # This might be better placed in a dedicated startup script or within FletAppView.__init__
    # if it's safe to call multiple times or if FletAppView is guaranteed to be a singleton.
    # For now, keeping it simple here.
    database_manager.initialize_database() # Assuming this function exists and sets up the DB
    ft.app(target=main)

# Note:
# - GuiController class is kept as is from the original file.
# - The main change is the introduction of FletAppView and moving tab definitions into its build method.
# - self.members_table_flet is created and placed in the "Membership Management" tab.
# - Other tabs retain their placeholders but are now part of the class structure.
# - The main function now instantiates FletAppView.
# - Added a call to database_manager.initialize_database() in if __name__ == "__main__":
#   This is a placeholder for where database initialization should occur.
#   The actual database_manager.py might need to be checked for the correct initialization function.
# - Corrected a few controller methods to properly return (bool, message) from database_manager calls.
# - Wrapped self.members_table_flet in an ft.Column as per example, though ft.DataTable can often be direct content.
#   This allows for adding more controls (like buttons above/below table) in that container later.
# - Adjusted expand properties for some placeholders in Column layouts as direct children of Column use `expand` differently.
#   Often, direct children of Column are sized intrinsically or use `height`/`width` unless also wrapped in `Container` with `expand`.
#   For simplicity, kept `expand` on containers.
# - The `database_manager` import might need an alias or path adjustment if the file structure is complex,
#   but `from reporter import database_manager` should work if `reporter` is the package root.
# - The `save_membership_action` and `delete_transaction_action` in `GuiController` were slightly adjusted to correctly handle the (bool, message) tuple from `database_manager`.
# - Fixed `generate_finance_report_excel_action` to ensure `df_columns` matches the data structure from `get_transactions_for_month`.
# - Removed redundant exception catch in `save_member_action` if `database_manager` handles its errors.
# - Ensured `toggle_plan_status_action` uses the message from `database_manager`.
# - Ensured `save_plan_action` correctly uses messages from `database_manager`.
# - Ensured `deactivate_member_action` uses message from `database_manager`.
# - Ensured `delete_plan_action` correctly returns message from `database_manager`.
# - Corrected `generate_custom_pending_renewals_action` and `generate_pending_renewals_action` to ensure they return an empty list for the third element of the tuple when no renewals are found, to match the type hint `list | None`.
# - Corrected `save_membership_action` to properly unpack the tuple from `database_manager.add_transaction` for both "Group Class" and "Personal Training" cases.
# - The `ft.Container` holding `self.members_table_flet` is set to `alignment=ft.alignment.top_center`. The table itself might also benefit from `expand=True` if it's the only child in the column and expected to fill. For now, `ft.Column([self.members_table_flet])` is used.
# - The `expand=True` on the `ft.DataTable` itself might be useful. Let's consider adding it directly to the table:
#   self.members_table_flet = ft.DataTable(..., expand=True)
#   and then the container could just be:
#   ft.Container(content=self.members_table_flet, expand=2, ...)
#   The current approach of ft.Column([self.members_table_flet]) is also fine.
#   The prompt specified "ft.Container(content=ft.Column([self.members_table_flet]) ... expand=2"
#   or "Or directly self.members_table_flet if it handles its own scrolling/sizing".
#   Let's stick to ft.Column([self.members_table_flet]) for now as it's more flexible for adding other controls later.
#   The outer container having expand=2 should make the column (and thus the table) fill the space.
#   If horizontal scrolling is needed, the DataTable itself has properties for that, or it can be wrapped in a Row.
#   The subtask mentions: "Add ft.Row([self.members_table_flet]) if the table needs to be scrollable horizontally later".
#   For now, ft.Column([self.members_table_flet]) is fine.

# Final check on GuiController.save_membership_action:
# It was returning `success` which was a tuple (bool, str) instead of just the boolean.
# Corrected it to return `True, "message"` or `False, db_message`.
# This was done by assigning `success, db_message = database_manager.add_transaction(...)`
# and then using `if success:`
# This has been applied.

    # Methods for Book Closing/Opening Dialogs and Actions MOVED to SettingsTab

    # Methods for Deleting Member, Plan, Transaction with Dialogs
    def on_delete_selected_member_click_flet(self, e):
        if self.selected_member_id_flet is None:
            self.member_actions_feedback_text.value = "No member selected to deactivate."
            self.member_actions_feedback_text.color = ft.colors.ORANGE
            self.member_actions_feedback_text.update()
            return

        member_name = "Unknown Member"
        # Find member name for confirmation from the table data
        # This assumes members_table_flet.rows contains DataRow objects with ft.Text in cells
        if self.members_table_flet and self.members_table_flet.rows:
            for row in self.members_table_flet.rows:
                if row.cells and len(row.cells) > 1 and isinstance(row.cells[0].content, ft.Text) and isinstance(row.cells[1].content, ft.Text):
                    try:
                        if int(row.cells[0].content.value) == self.selected_member_id_flet:
                            member_name = row.cells[1].content.value
                            break
                    except ValueError:
                        continue # Skip if cell content is not a valid int for ID

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Confirm Deactivate Member"),
            content=ft.Text(f"Are you sure you want to deactivate member '{member_name}' (ID: {self.selected_member_id_flet})? This member will be marked as inactive."),
            actions=[
                ft.TextButton("Yes, Deactivate", on_click=lambda ev: self._perform_delete_member_action(True, self.selected_member_id_flet)),
                ft.TextButton("No, Cancel", on_click=lambda ev: self._perform_delete_member_action(False, self.selected_member_id_flet)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        if self.page:
            self.page.dialog = dialog
            dialog.open = True
            self.page.update()

    def _perform_delete_member_action(self, confirmed: bool, member_id: int):
        if self.page.dialog:
            self.page.dialog.open = False
            self.page.update()

        if confirmed:
            success, message = self.controller.deactivate_member_action(member_id)
            self.member_actions_feedback_text.value = message
            self.member_actions_feedback_text.color = ft.colors.GREEN if success else ft.colors.RED
            if success:
                # Refresh member list (which will exclude inactive ones by default or show status)
                self.display_all_members_flet()
                # Refresh member dropdowns (to remove/update deactivated member)
                self.populate_member_dropdowns_flet()

                # If the deactivated member was the one selected, clear the specific history view
                if self.selected_member_id_flet == member_id:
                    self.selected_member_id_flet = None
                    # No direct members_table_flet.selected_index = None in Flet.
                    # Re-rendering the table via display_all_members_flet should visually clear selection
                    # if selection appearance depends on self.selected_member_id_flet.
                    self.display_membership_history_flet(None) # Clear specific member history
        else:
            self.member_actions_feedback_text.value = "Member deactivation cancelled."
            self.member_actions_feedback_text.color = ft.colors.ORANGE

        self.member_actions_feedback_text.update()
        if self.members_table_flet: self.members_table_flet.update() # Update table
        self.update() # General view update

    def on_delete_selected_plan_click_flet(self, e):
        if self.selected_plan_id_flet is None:
            self.plan_form_feedback_text.value = "No plan selected to delete."
            self.plan_form_feedback_text.color = ft.colors.ORANGE
            self.plan_form_feedback_text.update()
            return

        plan_name = "Unknown Plan"
        if self.plans_table_flet and self.plans_table_flet.rows:
            for row in self.plans_table_flet.rows:
                if row.cells and len(row.cells) > 1 and isinstance(row.cells[0].content, ft.Text) and isinstance(row.cells[1].content, ft.Text):
                    try:
                        if int(row.cells[0].content.value) == self.selected_plan_id_flet:
                            plan_name = row.cells[1].content.value
                            break
                    except ValueError:
                        continue

        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Confirm Delete Plan"),
            content=ft.Text(f"Are you sure you want to PERMANENTLY delete plan '{plan_name}' (ID: {self.selected_plan_id_flet})? This action cannot be undone and might affect historical records if not handled carefully by the backend."),
            actions=[
                ft.TextButton("Yes, Delete Plan", on_click=lambda ev: self._perform_delete_plan_action(True, self.selected_plan_id_flet)),
                ft.TextButton("No, Cancel", on_click=lambda ev: self._perform_delete_plan_action(False, self.selected_plan_id_flet)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        if self.page:
            self.page.dialog = dialog
            dialog.open = True
            self.page.update()

    def _perform_delete_plan_action(self, confirmed: bool, plan_id: int):
        if self.page.dialog:
            self.page.dialog.open = False
            self.page.update()

        if confirmed:
            success, message = self.controller.delete_plan_action(plan_id)
            self.plan_form_feedback_text.value = message
            self.plan_form_feedback_text.color = ft.colors.GREEN if success else ft.colors.RED
            if success:
                self.selected_plan_id_flet = None
                # self.plans_table_flet.selected_index = None # As above, no direct property
                self.display_all_plans_flet() # Refresh plans table
                self.populate_plan_dropdowns_flet() # Refresh plan dropdowns
                self.on_clear_plan_form_click(None) # Clear plan form
        else:
            self.plan_form_feedback_text.value = "Plan deletion cancelled."
            self.plan_form_feedback_text.color = ft.colors.ORANGE

        self.plan_form_feedback_text.update()
        if self.plans_table_flet: self.plans_table_flet.update()
        self.update()

    def on_delete_selected_transaction_click_flet(self, e):
        if self.selected_transaction_id_flet is None:
            self.history_actions_feedback_text.value = "No transaction selected to delete."
            self.history_actions_feedback_text.color = ft.colors.ORANGE
            self.history_actions_feedback_text.update()
            return

        # Transaction details for confirmation are not easily fetched without another controller call or complex table parsing.
        # Using just the ID is acceptable for now.
        dialog = ft.AlertDialog(
            modal=True,
            title=ft.Text("Confirm Delete Transaction"),
            content=ft.Text(f"Are you sure you want to PERMANENTLY delete transaction ID: {self.selected_transaction_id_flet}? This action cannot be undone."),
            actions=[
                ft.TextButton("Yes, Delete Transaction", on_click=lambda ev: self._perform_delete_transaction_action(True, self.selected_transaction_id_flet)),
                ft.TextButton("No, Cancel", on_click=lambda ev: self._perform_delete_transaction_action(False, self.selected_transaction_id_flet)),
            ],
            actions_alignment=ft.MainAxisAlignment.END,
        )
        if self.page:
            self.page.dialog = dialog
            dialog.open = True
            self.page.update()

    def _perform_delete_transaction_action(self, confirmed: bool, transaction_id: int):
        if self.page.dialog:
            self.page.dialog.open = False
            self.page.update()

        if confirmed:
            # Preserve current member selection if any, to refresh their specific history if needed
            member_id_for_specific_history_refresh = self.selected_member_id_flet

            success, message = self.controller.delete_transaction_action(transaction_id)
            self.history_actions_feedback_text.value = message
            self.history_actions_feedback_text.color = ft.colors.GREEN if success else ft.colors.RED
            if success:
                self.selected_transaction_id_flet = None
                # self.full_history_table_flet.selected_index = None # As above
                self.refresh_membership_history_display_flet() # Refresh full history table

                # If a member was selected and their specific history is shown, refresh it
                if member_id_for_specific_history_refresh is not None:
                    self.display_membership_history_flet(member_id_for_specific_history_refresh)
        else:
            self.history_actions_feedback_text.value = "Transaction deletion cancelled."
            self.history_actions_feedback_text.color = ft.colors.ORANGE

        self.history_actions_feedback_text.update()
        if self.full_history_table_flet: self.full_history_table_flet.update()
        self.update()

    def on_toggle_plan_status_click(self, e):
        """Handles the click event of the 'Toggle Active/Inactive' button."""
        if self.selected_plan_id_flet is None:
            self.plan_form_feedback_text.value = "Error: No plan selected to toggle status."
            self.plan_form_feedback_text.color = ft.colors.RED
            self.plan_form_feedback_text.update()
            return

        current_status_str = ""
        # Find the current status from the table
        for row in self.plans_table_flet.rows:
            if row.cells and len(row.cells) > 3 and isinstance(row.cells[0].content, ft.Text):
                try:
                    if int(row.cells[0].content.value) == self.selected_plan_id_flet:
                        if isinstance(row.cells[3].content, ft.Text):
                            current_status_str = row.cells[3].content.value # "Active" or "Inactive"
                            break
                except ValueError:
                    continue # Should not happen if IDs are integers

        if not current_status_str:
            self.plan_form_feedback_text.value = "Error: Could not determine current status of the selected plan."
            self.plan_form_feedback_text.color = ft.colors.RED
            self.plan_form_feedback_text.update()
            return

        current_status_bool = True if current_status_str == "Active" else False

        success, message, updated_plans = self.controller.toggle_plan_status_action(
            self.selected_plan_id_flet, current_status_bool
        )

        self.plan_form_feedback_text.value = message
        if success:
            self.plan_form_feedback_text.color = ft.colors.GREEN
            if updated_plans is not None:
                self.display_all_plans_flet(updated_plans)
            else: # Fallback if controller didn't return plans (should not happen on success)
                self.display_all_plans_flet()
            self.populate_plan_dropdowns_flet() # Refresh plan dropdowns
        else:
            self.plan_form_feedback_text.color = ft.colors.RED

        self.plan_form_feedback_text.update()
        self.update()


# Final check on GuiController.save_membership_action:
# It was returning `success` which was a tuple (bool, str) instead of just the boolean.
# Corrected it to return `True, "message"` or `False, db_message`.
# This was done by assigning `success, db_message = database_manager.add_transaction(...)`
# and then using `if success:`
# This has been applied.